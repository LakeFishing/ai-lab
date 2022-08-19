
##功能 :
## initial() 主要執行
## load_common_list() 讀取common_list，如果解題時，有遇到裡面的其中一個column元素，就要把那一個row全部都做判斷
## p4_create=> 建立dictionary(d1,d2) 
## solve()=> 解題，使用p4_questions裡面的function




#################################影響計算的參數#####################################################
# is_reverse = False  ##是否倒推計算
# is_minus = False  ##是否用原本數字下去扣
# is_plus = True   ##是否用一般計算，如果為False，且本身是+-，則要往下去找出=的項目做加減
# is_differ = False  ##是否用兩個數字相減
# is_total = False  ##判斷題目中是否有"一共"這類型詞彙，如果有，做倒推
# is_remain = False  ##判斷題目中是否有"剩下"這類型詞彙，如果有，做倒推
# change_plus = False   ##判斷題目中是否有"其中"這類型詞彙，如果有，做倒推
# plus_one = False  ##判斷是否要加1(當題目是排隊問題，問總排隊人數)
###############################影響計算的參數#####################################################


from doctest import FAIL_FAST
import sys

sys.dont_write_bytecode = True

from . import p4_create
from . import p4_questions
from openpyxl import load_workbook


def load_common_list():   ## 讀取common_list，如果解題時，有遇到裡面的其中一個column元素，就要把那一個row全部都做判斷

    wb = load_workbook(filename = 'project4/common_list.xlsx')
    sheet = wb.active
    common_list = {}

    
    col_num = 1  ##幾個column

    while True :
        col_num += 1 ##找出幾個column
        if (type(sheet.cell(row=1,column=col_num).value) !=str) and (type(sheet.cell(row=1,column=col_num).value) !=int):
            col_num -= 1 
            break
        
            

    for i in range(1,col_num+1):

        name = sheet.cell(row=1,column=i).value
        
        texts = []
        j = 2
        while True:
            
            if (type(sheet.cell(row=j,column=i).value) !=str) and (type(sheet.cell(row=j,column=i).value) !=int):
                break
            else:
                texts.append(sheet.cell(row=j,column=i).value)
                j+=1
                
        common_list[name]=texts   
        
            
            #q1.append(sheet.cell(row=1,column=i).value)   ##存問題之標籤名稱
            #q2.append(sheet.cell(row=2,column=i).value)   ##存問題之標籤種類


    
    return common_list

def solve(init_list,q1,q2,d1,d2,d3):  ##解題
    print()
    for i in d2:

        print(i,d2[i])



    name = ""  ##人名
    item = ""  ##物品
    items = []  ##存所有物品
    add=0  ##數量
    obj = "" ##接受者

    unit="" ##單位
    s_unit = "" ##另一個單位，做乘法用
    
    sum_ = 0  
    total = 0  ##最後輸出的答案(數字)
    quan = 0  ##問題中的數量，預設為0
    place = "" ##發生地
    adj = "" ##特點

    peoples = [] ##存入多個主事者

    m_num = 0##用來做乘除的數字

    plus_one = False  ##判斷是否要加1(當題目是排隊問題，問總排隊人數)
    
    who = ["誰","哪個","那個"] ##誰
    ask_who = False  ##問題是不是判斷"誰"最多或最少...
    
    back_ = ["原本","原來","原先","起初","一開始","剛開始","最早","最初","原有"] ##如果問題裡面有包含這些詞彙，則計算的正負是相反的(+=>- ， -=>+)

    is_reverse = False ##解題是否是倒推，例如原先有多少

    is_minus = False  ##解題是否用扣的

    boxs = ["他","她","它","他們","你","我"]  ##代名詞

    pos = [] ##人物在第幾句存在
    opr = ""  ##+-=

    unit_item={"錢":"元","位":"號"}  ##單位轉換

    common_list  = load_common_list()   ##存各種人(種類)

    places = []  ##存問題中的地點

    big = ""  ##大小(名稱) ex 快慢，遠近
    big_small = ""  ##紀錄詞性(大或小)

    s_plus = True ##是不是用一般的加減運算

    change_plus = False  ##判斷"="的項目是否全部用扣的

    add_back = False  ##判斷是否加回去

    no_have = False

    is_total = False ##問題中是否有"一共"

    first_total = False ##原本的句子中是否有"一共"

    is_sort = False  ##問題中是否包含"第"

    is_plus_one = False  ##判斷是否要將數字轉為quan

    enough = False  ##判斷夠不夠

    is_plus = True  ##是否用一般計算

    true_false = False ##是否用是非題計算

    sum_or_diff = ""  ##判斷是總和或相減

    oppo = {"大":"小",   ##相反詞dictionary => 主要用於question3 
            "小":"大",
            "年輕":"老",
            "老":"年輕",
            "遠":"近",
            "近":"遠",
            "多":"少",
            "少":"多",
            "長":"短",
            "短":"長",
            "高":"低",
            "低":"高",
            "新":"舊",
            "舊":"新",
            "快":"慢",
            "慢":"快",
            "左":"右",
            "右":"左",
            "早":"晚",
            "晚":"早",

            }

    minus_ = ["就","不夠","就有","還要","就可以","才","找回","才會","會","剪掉"]   ##用來判斷是否解題要用扣的
    differ = ["相差","差距","差"]  ##題目是否是用兩個數字做相減

    is_differ = False  ##判斷是否是用兩個數字做相減

    
    is_remain = False   ##判斷題目中是否有"剩下"，可以做倒推運算

    times = []  ##紀錄時間點標籤

    is_add = 0 ##判斷是否要額外加數字
    

    ##把問題的詞和標籤讀進來，做處理，並設定變數######################################################################
    
    for i in range(len(q1)):
        
        
        if q1[i]=="沒" :  ##沒投進
            
            is_reverse = True

        if q1[i] in ["便宜"] and q2[i]=="-":  ##如果有大或小的詞性，則指定big_small為此詞性，big為此詞彙
            
            change_plus = True

        if q2[i] in ["前"] :  ##如果有前或後
            
            is_plus_one = True

            if quan==0: ##如果quan=0，則把數字指定為1。ex : 我排第10個。我前面有幾人?
                
                quan = 1
                
        
        
        if q1[i] in back_:  ##判斷是否為倒推(原來,原本...)

            is_reverse = True

        if q1[i] in minus_ :   ##判斷是否用原本的數值下去扣
            
            is_minus = True

        
            
        elif q1[i]=="還":
            if i<len(q2)-2 and q1[i+1] in differ:  ##還差多少...
                
                is_minus = True   ##判斷是否用原本的數值下去扣
                
            elif i<len(q2)-2 and ('買' in q1 or '買了' in q1) and "剩下" in q1:
                
                is_minus = True   ##判斷是否用原本的數值下去扣

        if q1[i] in differ:   ##判斷是否用原本的數值下去扣


            is_differ = True


        if q2[i] =="特點" and q1[i] =="可以":   ##判斷是否用原本的數值下去扣

            true_false = True


        if q1[i]=="比" and q2[i]=="=":
            is_differ=True

            

        if q2[i] in ["=","+","-"]:
            
            
            opr = q1[i]
        
        
        if (q2[i]=="主事者" ) and q1[i] not in boxs:  ##如果是主事者，且不是代名詞，則主角為此主事者
            
            name=q1[i]

            peoples.append(name)

        elif (q2[i]=="主事者" ) and  q1[i] in boxs and  q1[i] in d2:

            name=q1[i]

            

            peoples.append(name)
            
        if q1[i] in who:  ##如果詞彙有在["誰","哪個","那個"]中，則可能是比大小的題目
                
            ask_who=True
        
                
        if q2[i] in ["=","-"]:

            if q2[i]=="-":
                sum_or_diff = q1[i]
            
        

            

        elif q2[i]=="物品" and q1[i]!="錢":##如果問題中有物品，item存入此物品(string，只會有一個)，items存入此物品(list，可能會有兩個)

          

            item=q1[i]
            items.append(q1[i])

            if "另" in item:  ##如果是"另一..."，則要做倒推
                change_plus=True

            if item in common_list:  ##如果有在common_list中，且d2有元素在common_list中，則把物品清空

                if item=="人":  ##如果物品為"人"，
                    for x in d2:
                        
                        if x in common_list[item]:
                            item = ""
                            break
                    if unit not in d2 and unit =="個":

                        unit = "人"
                        
                elif item in ["男生","女生"]:  ##如果物品為男生或女生
                    
                    check = 0
                    
                    for x in d2:

                        if x in common_list[item]:

                            check+=1

                    if check>=2:
                        
                        first_total=True  ##同"總共"，會做倒推
                        
                

        elif q2[i]=="單位" or (q1[i]=="錢" and q2[i]=="物品"):   ##找出單位，以供計算之比對

            
            #if unit=="":
                
            unit=q1[i]

##            elif s_unit=="":
##                
##                s_unit = unit ##存另一個單位，做乘除用
##                unit = q1[i]

            # if is_minus==True and unit!="" and s_unit!="":  ##如果是做減法(還、就可以...)，單位互換 => 不會用到

            #     x = unit
            #     unit = s_unit
            #     s_unit = x
                

            if unit=="人" and unit not in d2:
                
                unit="個"  ##如果單位為人，但沒有在d2，則將人改成個

           
                
                   
            
        elif q2[i]=="大" or q2[i]=="小":  ##如果有大或小的詞性，則指定big_small為此詞性，big為此詞彙
            big_small = q2[i]  ##大或小
            big=q1[i]

        
            

        elif q2[i]=="數量" :  ##判斷問題是否有數量
            
            try:
                quan=int(q1[i])
                if is_plus_one==True:
                    quan*=(-1)

            except:
                continue

        elif q2[i]=="接受者" :  ##判斷是否有接受者
           
            obj=q1[i]

            peoples.append(obj)

        elif q1[i] in ["共有","共","總共有","總共","一共"] :  ##判斷題目中是否有["共有","共","總共有","總共","一共"]這些詞彙
           
            is_total =True   

        elif q2[i]=="時間點" and q1[i] not in back_:  ##如果此時間點有在d2中，加入times中
            
            x = q1[i]
            if x in d2:
                
                times.append(x)
        

        elif q1[i]=="第" :##判斷有沒有"第"這個詞，如果有，表示可能是要做排名
           
            is_sort=True

        elif q2[i]=="發生地":

            place = q1[i]

            places.append(place)
        
            

        if "夠" and "不夠" in q1:
            enough = True
            true_false = True
            if is_minus==True and quan!=0:
                is_minus=False
                

        if i<len(q1)-2:
            
            if q1[i+1] in ["前","後"] and q1[i] in ['月','年','天','禮拜']:  ##判斷問題是否有某一時間點的前或後
               if q1[i+1]=="前":
                   is_add =-1
               
                
               elif q1[i+1]=="後":
                   is_add =1
            elif q2[i] in ["時間點"] and q1[i] in ["去年"]:
                
                is_add =-1
                quan = 1

##    if s_unit==unit:  ##如果兩個單位一樣，則把第二單位拿掉
##        
##        s_unit=""
##        
##
##
##    if s_unit!="" and unit!="" and s_unit!=unit:  ##如果確定有兩個單位不同，可能會做乘法
##
##        try:
##            n = q1.index(s_unit)
##
##            m_num = int(q1[n-1])   ##找出拿來乘除的數字
##
##            
##            
##        except:
##            print()
    
    if name=='' and len(times)==1:
        name = times[0]
    
               
    if len(items)==3:
        del items[0]##如果物品數量有三個，則刪掉第一個


    if unit not in d2:  ##如果單位沒有出現在d2中，則要從d2找出單位

        if name!="" and name in d2:
            
            x = d2[name][0][0]  ##找到主事者最底層位置(單位)   
            unit  = d1[x][0]
            
        

    if unit in unit_item and unit not in d2: ##如果單位沒有在題目中，但是有在轉換單位的dictionary，則單位轉換
        unit = unit_item[unit]
              
    
    
    if "比較" in d2  :
        is_plus = False  ##比較，本身是等於的項目，不做加減，如果本身是+-，則要往下去找出=的項目做加減

    if "比" in q1:  ##如果"比"在問題中，則會做特殊運算
        is_plus = False

    if "總共" in d2:  ##如果總共在d2中，first_total=true，會做倒推運算
        first_total = True

    if "剩下" in d2:   ##如果剩下在d2中，則會做倒推運算
        is_remain = True

    if "其中" in d2:

        change_plus = True

    if "總排" in d2:
        plus_one = True

    if "不夠" in d2:
        add_back = True
    
    if "沒" in d2:
        no_have = True

    if "沒有" in d2:
        no_have = True

    if "沒" in q1:
        no_have = True

    if "沒有" in q1:
        no_have = True

    if "完" in d2:
        no_have = True

    
    
    p4_questions.attr(is_reverse,is_minus,is_plus,is_differ,is_total,is_remain,change_plus,quan,m_num,s_unit,add_back,no_have)   ##設定解題時的變數

    
    if name!="" and len(times)==1 and times[0] in d2 :   ##如果有時間，又有主事者，則時間會變成主事者
        if d2[times[0]][3][0]!="none":
            name = times[0]
        
        
        
    elif len(times)==2 and times[0] in d2 and times[1] in d2 and (is_differ==True or is_plus==False):  ##如果有時間，又有主事者接受者，則時間會變成主事者接受者
        if name!="":
            name = ""
        if item!="" and len(items)>0:
            items.clear()
            item = ""
        #items.append(times[0])
        #items.append(times[1])

    
    if name=="" and place!="":  ##如果沒有主事者，則將發生地改成主事者
        if place in d2:
            if d2[place][1][0]!="none":
                name = place


#####################################解題###############################################################
    
    if is_sort ==True: ##如果問題裡面有"第"

        total = p4_questions.question10(name,unit,d1,d2,d3,quan)
        #print()
        #print("解題")
        #print(q1)
        
        #print("答案:",item,total,unit)

    elif (len(places)>1 and item=="") or (len(places)>1 and item=="距離"):   ##如果試問兩個發生地的距離

        total = p4_questions.question11(name,places,unit,d1,d2,d3,quan)
        
    
    elif first_total==True: ##如果原本句子裡面有"總共"

        
        if unit in unit_item and unit not in d2:
            unit = unit_item[unit]
        total = p4_questions.question8(name,item,unit,d1,d2,d3,quan)
        ##問數量(問題中，包含主事者,物品,單位...，原本句子裡面有'總共'
        
        #print("解題")
        #print(q1)
        
        #print("答案:",item,total,unit)
        

    elif ((name!="" and name in d2 and  obj !="" and obj in d2) and (item!="" or unit!="")) or (len(peoples)>=2) :   ##如果問題中，包含主事者,接受者,物品,單位，問差距
        #print("1.人名:",name," ",obj," ""2.物品:",item,"3.單位:",unit)
        people = []
        if name in d2:
            people = d2[name][0]    ##輸入名稱，找出名稱在那些句子存在，例如: 輸入"我"，存在於a,b
        
        if unit in unit_item and unit not in d2:
            unit = unit_item[unit]
        if big_small =='':
            total = p4_questions.question6(people,obj,name,item,unit,d1,d2,d3,quan,peoples)
        elif big_small !='':

            total = p4_questions.question3(init_list,name,item,unit,d1,d2,d3,big,big_small,oppo,true_false)

        if is_add==1 and quan!=0:  ##判斷是否要用問題中的數字再做加減
            total+=quan
        elif is_add==-1 and quan!=0:
            total-=quan
            
        #print()
        #print("解題")
        #print(q1)
        
        #print("答案:",item,total,unit)

    elif len(items)==2 and unit!="" :

        
        
        a = items[0]
        b = items[1]

        ans_list = []
        

        if a not in d2 or b not in d2:  ##如果其中一個物品不在d2中，則用question5去解

            

            if a in common_list:
                for i in d2:
                    if i in common_list[a] and i not in ans_list:
                        ans_list.append(i)
            elif b in  common_list:
                for i in d2:
                    if i in common_list[b] and i not in ans_list:
                        ans_list.append(i)

            if len(ans_list)!=0:
                
                if a in d2 and len(d2[a][1]) and d2[a][1][0]!='none':
                    for i in (ans_list):
                        people = d2[a][0]
                        
                        total += p4_questions.question1(people,a,i,unit,d1,d2,d3,quan,s_unit,m_num)

                        
                else:
                    for i in (ans_list):
                        
                        total += p4_questions.question2(i,unit,d1,d2,d3,quan,s_unit,m_num)

            else:

                total = p4_questions.question5(unit,d1,d2,d3,quan)
            #print()
            #print("解題")
            #print(q1)

            if is_add==1 and quan!=0:  ##判斷是否要用問題中的數字再做加減
                total+=quan
            elif is_add==-1 and quan!=0:
                total-=quan
                
            #print("答案:",name,items,total,unit)

        elif d2[a][3][0]!="none" :  ##如果第一個物品的數量不為none，視為主事者
            
            
            #total = p4_questions.question7(items,unit,d1,d2,d3,is_reverse,quan,is_minus,is_differ,is_plus,is_total)

            if d2[b][3][0]=="none":
                people = d2[a][0]

                if big_small!="":
                    total = p4_questions.question3(init_list,name,item,unit,d1,d2,d3,big,big_small,oppo,true_false)
                    return "","","","",total

                total = p4_questions.question1(people,a,b,unit,d1,d2,d3,quan,s_unit,m_num)

            else:
                
                for i in (items):
                    
                    total += p4_questions.question2(i,unit,d1,d2,d3,quan,s_unit,m_num)
                    
            
            if is_add==1 and quan!=0:  ##判斷是否要用問題中的數字再做加減
                total+=quan
            elif is_add==-1 and quan!=0:
                total-=quan
                
            #print()
            #print("解題")
            #print(q1)
            
            #print("答案:",name,items,total,unit)
        
        elif d2[b][3][0]!="none":
            total = p4_questions.question2(b,unit,d1,d2,d3,quan,s_unit,m_num)
            
        else:
            
            total = p4_questions.question9(items,unit,d1,d2,d3,quan)
            
            if is_add==1 and quan!=0:  ##判斷是否要用問題中的數字再做加減
                total+=quan
            elif is_add==-1 and quan!=0:
                total-=quan
                
            #print("答案:",name,items,total,unit)   ##問題中包含兩物品(兩個物品數量為none)
    
    
    elif name!="" and item!="" and unit!="" and name in d2 and item in d2 :   ##如果問題中，包含主事者,物品,單位

        
        #print("1.人名:",name,"2.物品:",item,"3.單位:",unit)
        people = d2[name][0]    ##輸入名稱，找出名稱在那些句子存在，例如: 輸入"我"，存在於a,b
        if unit in unit_item and unit not in d2:
            unit = unit_item[unit]

        if item not in d2 and item in common_list:   ##如果物品不在d2，且在common_list，則要一個一個算出
            ans_list = []

            for i in common_list[item]:
                if i in d2:
                    ans_list.append(i)
            
            if len(ans_list)!=0:
                    
                for i in (ans_list):
                    
                    total += p4_questions.question1(people,name,i,unit,d1,d2,d3,quan,s_unit,m_num)

        else:
            
            total = p4_questions.question1(people,name,item,unit,d1,d2,d3,quan,s_unit,m_num)

        if is_add==1 and quan!=0:  ##判斷是否要用問題中的數字再做加減
            total+=quan
        elif is_add==-1 and quan!=0:
            total-=quan

        
        #print()
        #print("解題")
        #print(q1)
        
        #print("答案:",name,item,total,unit)

    

    

    elif ask_who==True or big!="":  ##比大小
        #print("1.人名:",name,"2.物品:",item,"3.單位:",unit)
        total = p4_questions.question3(init_list,name,item,unit,d1,d2,d3,big,big_small,oppo,true_false)
        print()
        #print("解題")
        #print(q1)
        unit = ""

        item = ""
        #print("答案:",total)

    
        return "","","","",total

    elif name!="" and (item!="" or unit!="") and ask_who!=True and name in d2:   ##如果問題中，包含主事者,單位
        #print("1.人名:",name,"2.物品:",item,"3.單位:",unit)
        people = d2[name][0]    ##輸入名稱，找出名稱在那些句子存在，例如: 輸入"我"，存在於a,b
        
        if unit in unit_item and unit not in d2:
            unit = unit_item[unit]

        if unit=="樓" and "層" in d2:
           
            
            total = p4_questions.question4(people,name,item,unit,d1,d2,d3,quan,s_unit,m_num)
            total +=p4_questions.question4(people,name,item,"層",d1,d2,d3,quan,s_unit,m_num)

            
        elif unit=="歲" and "年" in d2:

            total = p4_questions.question4(people,name,item,unit,d1,d2,d3,quan,s_unit,m_num)
            total +=p4_questions.question4(people,name,item,"年",d1,d2,d3,quan,s_unit,m_num)

        else:
            print(is_add)
            total = p4_questions.question4(people,name,item,unit,d1,d2,d3,quan,s_unit,m_num)

        if is_add==1 and quan!=0:  ##判斷是否要用問題中的數字再做加減
            total+=quan
        elif is_add==-1 and quan!=0:
            total-=quan

        print()
        #print("解題")
        #print(q1)
        
        #print("答案:",name,item,total,unit)

    
    elif name=="" and item!="" and unit!="" and (ask_who!=True or big=="") and (item in d2 or item in common_list):  ##如果問題中，包含物品,單位
        
        #print("1.人名:",name,"2.物品:",item,"3.單位:",unit)
        if unit in unit_item and unit not in d2:
            unit = unit_item[unit]

        if item not in d2 and item in common_list:
            ans_list = []
            

            for i in common_list[item]:
                if i in d2:
                    ans_list.append(i)
            
            if len(ans_list)!=0:
 
                for i in (ans_list):
                    
                    total += p4_questions.question2(i,unit,d1,d2,d3,quan,s_unit,m_num)
            else:
                total = p4_questions.question5(unit,d1,d2,d3,quan)
            
        else:
            
            total = p4_questions.question2(item,unit,d1,d2,d3,quan,s_unit,m_num)
            

        if is_add==1 and quan!=0:  ##判斷是否要用問題中的數字再做加減
            total+=quan
        elif is_add==-1 and quan!=0:
            total-=quan
        elif is_plus_one==True:  ##前面排幾個
            total-=quan
                
        print()
        #print("解題")
        print(q1)
        
        #print("答案:",name,item,total,unit)


    elif (unit!="" and unit in d2) :
        
        
        print("1.人名:",name,"2.物品:",item,"3.單位:",unit)

        if unit=="人" and "個" in d2:
            
            total = p4_questions.question5(unit,d1,d2,d3,quan)

            if change_plus == True:
                total = abs(total-p4_questions.question5("個",d1,d2,d3,quan))
            else:
                total += p4_questions.question5("個",d1,d2,d3,quan)

        
        
        elif unit=="樓" and "層" in d2:
            
            total = p4_questions.question5("樓",d1,d2,d3,quan)
            
            total += p4_questions.question5("層",d1,d2,d3,quan)
                     
        else:
            total = p4_questions.question5(unit,d1,d2,d3,quan)

        

        if is_add==1 and quan!=0:  ##判斷是否要用問題中的數字再做加減
            total+=quan
        elif is_add==-1 and quan!=0:
            total-=quan

        
        print()
        #print("解題")
        #print(q1) 
        #print("答案:",total,unit)

    elif item!="" and unit=="":

        
        
        total = p4_questions.question2(item,unit,d1,d2,d3,quan,s_unit,m_num) 
        ##如果問題中不包含單位

        if is_add==1 and quan!=0:  ##判斷是否要用問題中的數字再做加減
            total+=quan
        elif is_add==-1 and quan!=0:
            total-=quan


        print()
        #print("解題")
        #print(q1)
        
        #print("答案:",name,item,total,unit)
        
        
    else:
        if true_false==True:
            total =p4_questions.question12(name,item,unit,d1,d2)
            return "","",total,"",""
        else:
            print("題目有誤")

    

    if enough==True and quan!=0:

        if quan>0 and quan>total:
            
            return "","","不夠","",""
        
        else:
            
            return "","","夠","",""

    if obj!="":
        name=name+"和"+obj+sum_or_diff
    if len(peoples)>1:
        name = ""
        for i in peoples:
            name = name+i+","

    
    if name=="" and item!="":   #如果主事者為空，但物品不為空，則物品視為主事者

        name,item = item,name

    if plus_one==True:
        total +=1

    if len(items)>1:

        item = ""
        if name!="":
            name = ""
        for i in items:

            item+=i+","
        
        
    
    return name,opr,item,total,unit
    
def initial():  ##讀檔


    
    f = open('project3/word.txt','r',encoding='utf-8')   ##開啟資料結構
    q = open('project3/question.txt','r',encoding='utf-8') ##開啟題目
    
    line = f.readlines()
    line2 = q.readlines()
    
    
    
    init_list,q1,q2,d1,d2,d3 = p4_create.create_table(line,line2)  ##讀取資料結構，建立dictionary

    name,opr,item,total,unit = solve(init_list,q1,q2,d1,d2,d3)  ##解題

        
    #ans = name+""+opr+""+item+""+str(total)+""+unit
    ans = str(total)+""+unit

    f.close()
    q.close()

    return ans


# ans = initial()  ##執行

# print('本題答案：' + ans)

