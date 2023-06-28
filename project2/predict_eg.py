import openpyxl
import numpy as np

import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

from keras.models import load_model

model01 = load_model('project2/t1h5/CNN01.h5')
model02 = load_model('project2/t1h5/CNN02.h5')
model03 = load_model('project2/t1h5/CNN03.h5')
model04 = load_model('project2/t1h5/CNN04.h5')
model05 = load_model('project2/t1h5/CNN05.h5')
model06 = load_model('project2/t1h5/CNN06.h5')
model07 = load_model('project2/t1h5/CNN07.h5')
model08 = load_model('project2/t1h5/CNN08.h5')
model09 = load_model('project2/t1h5/CNN09.h5')
model10 = load_model('project2/t1h5/CNN10.h5')
model11 = load_model('project2/t1h5/CNN11.h5')
model12 = load_model('project2/t1h5/CNN12.h5')
model13 = load_model('project2/t1h5/CNN13.h5')

def predict_eg():
    d_tag = {0: '主事者', 1: '接受者', 2: '=', 3: '+', 4: '-',
             5: '數量', 6: '單位', 7: '特點', 8: '物品', 9: '虛詞',
             10: '時間點', 11: '發生地', 12: '大', 13: '小',
             14: '連接詞', 15: '前', 16: '後'}
    fn = 'project1/process/proj1.xlsx'
    fn1 = 'proj2.xlsx'
    wb = openpyxl.load_workbook(fn)
    ws = wb['Sheet2']

    li = []
    for i in range(1, ws.max_row + 1):
        li.clear()
        wb = openpyxl.load_workbook(fn)
        ws = wb['Sheet2']
        rowA = ws[i]
        inputData = [rowA[x].value for x in range(len(rowA))]
        length = sum(inputData)

        inputData = np.array(inputData)
        inputData = inputData.reshape(1, 32, 32, 1)
        inputData = inputData.astype('float32')
        pred01 = model01.predict(inputData)
        li.append(pred01.argmax())

        if length >= 2:
            pred02 = model02.predict(inputData)
            li.append(pred02.argmax())
        if length >= 3:
            pred03 = model03.predict(inputData)
            li.append(pred03.argmax())
        if length >= 4:
            pred04 = model04.predict(inputData)
            li.append(pred04.argmax())
        if length >= 5:
            pred05 = model05.predict(inputData)
            li.append(pred05.argmax())
        if length >= 6:
            pred06 = model06.predict(inputData)
            li.append(pred06.argmax())
        if length >= 7:
            pred07 = model07.predict(inputData)
            li.append(pred07.argmax())
        if length >= 8:
            pred08 = model08.predict(inputData)
            li.append(pred08.argmax())
        if length >= 9:
            pred09 = model09.predict(inputData)
            li.append(pred09.argmax())
        if length >= 10:
            pred10 = model10.predict(inputData)
            li.append(pred10.argmax())
        if length >= 11:
            pred11 = model11.predict(inputData)
            li.append(pred11.argmax())
        if length >= 12:
            pred12 = model12.predict(inputData)
            li.append(pred12.argmax())
        if length >= 13:
            pred13 = model13.predict(inputData)
            li.append(pred13.argmax())

        excel_save = list(map(d_tag.get, li))
        wb = openpyxl.load_workbook(fn1)
        ws = wb.active
        for j in range(len(excel_save)):
            ws.cell(row = 2 + 2 * i, column = j + 1, value = excel_save[j])
        wb.save('proj2.xlsx')

    print('done predict_eg.py')

if __name__ == '__main__':
    predict_eg()