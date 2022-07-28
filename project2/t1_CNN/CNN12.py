from keras.layers import Dense, Conv2D, Flatten, MaxPooling2D
from keras.models import Sequential
import tensorflow as tf
import numpy as np
import openpyxl
import sys

fn = "makedata/output_100_DB.xlsx"
wb = openpyxl.load_workbook(fn)
ws = wb.worksheets[1]

row_list = []
for i in range(1, ws.max_row + 1):
    row = ws[i]
    row_list_t = [row[x].value for x in range(len(row))]
    if sum(row_list_t) > 11:
        row_list.append(row_list_t)
ws = wb.worksheets[2]

num_list = []
for j in range(1, ws.max_row + 1):
    num = ws[j]
    num_list_t = [num[x].value for x in range(len(num))]
    while None in num_list_t:
        num_list_t.remove(None)
    num_list.append(num_list_t)

otn_list = []
for otn in range(1, ws.max_row + 1):
    if len(num_list[otn - 1]) > 11:
        otn_list_t = num_list[otn - 1][11]
        otn_list.append(otn_list_t)

if otn_list == []:
    sys.exit()

x_train = np.array(row_list)
y_train = np.array(otn_list)
x_train = x_train.reshape(len(otn_list), 28, 28, 1)
y_train = y_train.reshape(len(otn_list))
x_train = x_train.astype('float32')
y_train = y_train.astype('uint8')

x_test = np.array(row_list)
y_test = np.array(otn_list)
x_test = x_test.reshape(len(otn_list), 28, 28, 1)
y_test = y_test.reshape(len(otn_list))
x_test = x_test.astype('float32')
y_test = y_test.astype('uint8')
input_shape = (28, 28, 1)

print('here')

model = Sequential()
model.add(Conv2D(28, kernel_size=(2, 2), input_shape=input_shape))
model.add(MaxPooling2D(pool_size=(2, 2)))
model.add(Flatten())
model.add(Dense(128, activation=tf.nn.relu))
model.add(Dense(32, activation=tf.nn.relu))
model.add(Dense(17, activation=tf.nn.softmax))

model.compile(optimizer = 'adam',
              loss = 'sparse_categorical_crossentropy',
              metrics = ['accuracy'])
model.fit(x = x_train, y = y_train, epochs = 100)

scores = model.evaluate(x_test, y_test)
print('scores', scores)

# model.save('./t1h5/CNN12.h5')
model.save('./test/t1h5/CNN12.h5')
del model