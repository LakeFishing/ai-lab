from tkinter import N
import openpyxl
import numpy as np


def proj1():
    fn = 'project1/process/斷詞.xlsx'
    wb = openpyxl.load_workbook(fn)
    wb.create_sheet(index=1, title='Sheet2')
    wb.create_sheet(index=2, title='Sheet3')
    ws = wb['Sheet1']
    d = {'v': '0', 'pv': '1', 'nv': '2', 'n': '3', 'u': '4',
         'adv': '5', 'int': '6', 'adj': '7', 'name': '8', 'time': '9',
         'equal': '10', 'place': '11', 'range': '12', 'total': '13',
         'special': '14', 'passive': '15', 'bigger': '16', 'smaller': '17',
         'after': '18', 'before': '19', 'p': '20', 'add': '21', 'sub': '22', 'and': '23',
         'dunit': '24', 'bunit': '25', 'cunit': '26', 'tunit': '27',
         }

    colA = ws['A']
    colA_list = [colA[x].value for x in range(len(colA))]
    colA_list = [x for x in colA_list if x is not None]
    temp = 1
    while '、' in colA_list:
        ws_active = wb.active
        ws_active['B' + str(colA_list.index('、') + temp)] = 'dddd'
        wb.save('project1/process/斷詞.xlsx')
        colA_list.remove('、')
        temp += 1

    ja = ' '.join(colA_list)
    sa = ja.split('，')
    sen_A = []
    for i in range(len(sa)):
        l = sa[i].split(' ')
        while '' in l:
            l.remove('')
        while '\n' in l:
            l.remove('\n')
        sen_A.append(l)

    colB= ws['B']  #同A欄舉動處理B欄類別
    colB_list=[colB[x].value for x in range(len(colB))]
    colB_list=[x for x in colB_list if x is not None]
    jb=' '.join(colB_list)  
    sb=jb.split('x')  
    sen_B=[]
    for i in range(len(sb)):  #print(s[i])
        l=sb[i].split(' ')    #print(l)
        while '' in l:
            l.remove('')  #print(l)
        while 'dddd' in l:
            l.remove('dddd')  #print(l)
        sen_B.append(l)
        sen_B_alter= [x for x in sen_B if x != []]
#    print(sen_B_alter)

    for j in range(len(sen_A)):  #處理每一句話
        colA_word=sen_A[j]  #取一句話:第一句話、第二句話...
        #print(colA_word)
        colA_position=[]
        for i, element in enumerate(colA_word):  #A欄斷詞對應斷詞位置，得目標
            colA_position.append(i)
        #print(colA_position) #[0, 1, 2, 3]=colA
        
        colB_word=sen_B_alter[j]  #處理每一句話
        colB_position_w=list(map(d.get,colB_word))  #B欄類別分類對應到類別位置，參考字典檔
#        print(colB_position_w)
        colB_position=list(map(int, colB_position_w))  #字串轉數值
#        print(colB_position) #[13, 14, 12, 1]=colB
        
        if len(colB_position)!= len(colA_position):
            del colA_position[-1]
        #print(colA_position)
        
        #input矩陣25x25
        t_input=np.zeros([28,28])
        
        for i in range(len(colA_position)):  #(col1,col2)=1填入矩陣
            t_input[colB_position[i]][colA_position[i]]=1
        #print(t_input) #二維矩陣
        t1_input=t_input.flatten()  #矩陣轉一維
        t2_input=t1_input.tolist()  #numpy是array，轉list  #print(t2_input)
        
        #ws_active=wb.active  #確定目前開啟的是哪張工作表
        #print('目前工作表名稱=',ws_active.title)
        ws2= wb["Sheet2"] #print("目前工作表=", ws.title)
        #ws['A1']='python'
        ws2.append(t2_input) #寫入處理的話



    #同理
    #colE=[1,2,3,4]  =求解問題斷詞
    #colF=[16,4,2,6]  =每個斷詞對應到的類別分類
        
    colE= ws['E']  #讀E欄，目前將求解視為一句話處理
    colE_list=[]
    for x in range(len(colE)):
        if not colE[x].value == None:
            colE_list.append(colE[x].value)  #用list存E欄每一格數值
#    print(colE_list) #['剩下', '幾', '個', '飯糰']
    tempE=1    
    while '、' in colE_list:
        ws_active=wb.active 
        ws_active['F'+str(colE_list.index('、')+tempE)]='dddd'
        wb.save('project1/process/斷詞.xlsx')
        colE_list.remove('、')
        tempE+=1
        
    colE_position=[]
    for k, element in enumerate(colE_list):  #E欄斷詞對應斷詞位置，得目標
        colE_position.append(k)
#    print(colE_position)  #[0, 1, 2, 3]=colE



    colF= ws['F']  #同E欄舉動處理F欄類別
    colF_list=[] 
    #colF_list=[colF[x].value for x in range(len(colF))] 
#    print(colF_list)  #['total', 'adv', 'unit', 'n', None, None, None, None, None, None, None]
    for x in range(len(colF)):
        if not colF[x].value == None:
            colF_list.append(colF[x].value)  #用list存F欄每一格數值
    #print(colF_list) #['total', 'adv', 'unit', 'n']
    
    if 'x' in colF_list:
        colF_list.remove('x')
    while 'dddd' in colF_list:
        colF_list.remove('dddd')  #print(l)    
#    print(colF_list)
    
    
    colF_position_w=list(map(d.get,colF_list))  #B欄類別分類對應到類別位置，參考字典檔
    #解決excel中出現x的問題    
    if len(colE_position) > len(colF_position_w):
        length = len(colE_position)-len(colF_position_w)
        del colE_position[-length]
    else:
        if len(colF_position_w) > len(colE_position):
            length = len(colF_position_w)-len(colE_position)
            del colF_position_w[-length]
    if None in colF_position_w:
        colF_position_w.remove(None)
    
    colF_position=list(map(int, colF_position_w))  #字串轉數值
#    print(colF_position) #[15, 3, 1, 5]=colF
    

    #input矩陣25x25
    tr_input=np.zeros([28,28])
    for r in range(len(colF_position)):  #(colE,colF)=1填入矩陣
        tr_input[colF_position[r]][colE_position[r]]=1
    #print(tr_input) #二維矩陣
    tr1_input=tr_input.flatten()  #矩陣轉一維
    tr2_input=tr1_input.tolist()  #numpy是array，轉list  #print(tr2_input)

    #ws_active=wb.active  #確定目前開啟的是哪張工作表
    #print('目前工作表名稱=',ws_active.title)
    ws3= wb["Sheet3"] #print("目前工作表=", ws.title)
    #ws['A1']='python'
    ws3.append(tr2_input) #寫入處理的話

    #wb.save('input.xlsx')
    wb.save('project1/process/proj1.xlsx')

    wb= openpyxl.Workbook() #開新檔案
    ws= wb['Sheet']
    # wb1 = openpyxl.load_workbook('project1/output/ans.xlsx')
    # ws1 = wb1['Sheet']
    for i in range(len(sen_A)):
        for j in range(len(sen_A[i])):
    #        print(3+2*i,1+j)
    #        print(sen_A[i][j])
            ws.cell(row=3+2*i,column=1+j,value=sen_A[i][j])
            # ws1.cell(row=3+2*i,column=1+j,value=sen_A[i][j])
    for i in range(len(colE_list)):
        ws.cell(row=1,column=1+i,value=colE_list[i])
        # ws1.cell(row=1,column=1+i,value=colE_list[i])

    wb.save("proj2.xlsx")
    # wb1.save('project1/output/ans.xlsx')
    print('done proj1.py')

if __name__ == '__main__':
    proj1()