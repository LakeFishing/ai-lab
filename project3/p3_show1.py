##開新的視窗，在canvas上畫出命題語意網路，以及存成jpg檔

##使用function: draw(cv,sheet) : 

### 邏輯:

### 1.在每一個句子中，先找出 " +-=大小前後 " 運算符號，接下來以此為"基準"繪圖

### 2.在基準之左邊的標籤，就畫在其左邊;在基準右邊的標籤，就畫在其右邊


### 範例:

##我有15個水果，你給我3個水果，我共有多少水果?

### 標籤名稱:我,=,15,個,水果，你,給,我,3,個,水果

### 從標籤中可找出"="，因此，=左邊的標籤有"我"，因此"我"會畫在"="的左邊，而剩下的標籤都畫在"="的右邊



from tkinter import *
from random import randint
import pyautogui
from openpyxl import load_workbook

oval_w = 60 #橢圓的寬
oval_h = 30   #橢圓的高


def draw(cv,sheet):       ###命題語意網路
    
    print()
    print("命題語意網路")

    ###繪圖邏輯:
    ###在每一個句子中，先找出 " +-=大小前後 " 運算符號，接下來以此為"基準"繪圖
    ###在基準之左邊的標籤，就畫在其左邊;在基準右邊的標籤，就畫在其右邊
    
    ###範例:我有15個水果，你給我3個水果，我共有多少水果?
    ###標籤名稱:我,=,15,個,水果，你,給,我,3,個,水果
    ###從標籤中可找出"="，因此，=左邊的標籤有"我"，因此"我"會畫在"="的左邊，而剩下的標籤都畫在"="的右邊
    

    keys = ["+","-","=","大","小","前","後"]    ##運算符號，供之後判斷參考用
    
    #預設row及column 存放供畫圖計算用，如果標籤為+-=大小前後，則存進去。之後以這裡面的元素為基底，進行繪圖
    cols = []
    rows = []
    #預設row及column 存放供畫圖計算用，如果標籤為+-=大小前後，則存進去。之後以這裡面的元素為基底，進行繪圖

    

    ##起始座標
    init_x = 300  
    init_y = 50
    ##起始座標

    ##起始座標增加幅度
    addx = 0
    addy = 120
    ##起始座標增加幅度

    ##下一個起始座標
    next_x = init_x
    next_y = init_y
    ##下一個起始座標

    max_x = 0 ##紀錄寬度最大值
    
    count = 1  #計算有幾個句子
    
    items = 0 ##計算有多少個標籤，如果超過7個標籤，下一句話就往下面去畫，不往旁邊

    is_plus = 0
    


    for i in range(4,100,2):   ##找出每一句話的中間點，找出每一句話+-=大小前後，將其視為中間點 => excel中的row=這一句話

        is_plus = 0
        
        if (type(sheet.cell(row=i,column=1).value) !=str) and (type(sheet.cell(row=i,column=1).value) !=int) :  ##如果資料為空的，就退出
            break

        

        for j in range(1,100):  ##   => excel中的column，這句話的第幾個字
            
            if (type(sheet.cell(row=i,column=j).value) !=str) and (type(sheet.cell(row=i,column=1).value) !=int) :  ##如果資料為空的，就退出
                if is_plus == 0:   #如果沒有+-=大小前後，則存進去中間那個標籤
                    rows.append(i)        
                    cols.append(int(j/2))
                
                break
            else:
                if sheet.cell(row=i,column=j).value in keys:
                    rows.append(i)        
                    cols.append(j)                   #如果標籤為+-=大小前後，則存進去
                    is_plus = sheet.cell(row=i,column=j).value
                    break
    
    
    
    print("運算符號在每句話存在的位置:(column)",cols)
    print()
    
    
    for i in range(len(cols)):  ##前面跑完後，得到每一句話的col(中間點)，根據中間點的前後來繪圖
    
        print("第",str(i+1),"句話","，中間點的column為",cols[i])

        items = 0  ##計算畫到第幾個字
        
        current_row = rows[i]
        next_x = init_x
        
        a = cols[i]            ##把column的標籤分兩邊去畫
        b =  cols[i]+1


        ##顯示為第幾個句子
        cv.create_oval(init_x-oval_w/2,init_y-oval_h/2,init_x+oval_w/2,init_y+oval_h/2,fill="white")     
        cv.create_text(init_x,init_y,text="句子"+str(count),font=("標楷體",10))   
        ##顯示為第幾個句子
        
        
        print("畫左邊")
        for j in range(a,0,-1):   ##畫左邊
            
            cv.create_line(init_x,init_y+oval_h/2,init_x+addx,init_y+addy)   ##先畫線
            
            cv.create_text(init_x+addx-10,init_y+addy-oval_h/2-10,text=sheet.cell(row=current_row,column=j).value,font=("標楷體",10))   ##上文字(標籤)

            cv.create_oval(init_x+addx-oval_w/2,init_y+addy-oval_h/2,init_x+addx+oval_w/2,init_y+addy+oval_h/2,fill="white")  ##畫橢圓
            
            cv.create_text(init_x+addx,init_y+addy,text=sheet.cell(row=current_row-1,column=j).value,font=("標楷體",10))  ##上文字(詞彙)
            
            print(sheet.cell(row=current_row-1,column=j).value,"，座標:",str(init_x+addx),str(init_y+addy))

            addy-=10
            if(init_y+addy<init_y):
                addx+=80
            else:
                addx-=80
            items+=1
            
        ##重設座標增加幅度
        addx = 0    
        addy = 120
        ##重設座標增加幅度
        
        
        print("畫右邊")
        for k in range(b,11):    #畫右邊

            ##重設座標增加幅度
            addy-=10

            if(init_y+addy<init_y):
                addx-=80
            else:
                addx+=80

            ##重設座標增加幅度

                
            ##設定下一個起始座標
            if (init_x >= next_x):
                next_x = init_x

            if(init_y+addy>=next_y):
                next_y = init_y+addy
             ##設定下一個起始座標

                
            if (type(sheet.cell(row=current_row,column=k).value) !=str) :
                break


            
            else:
                
                cv.create_line(init_x,init_y+oval_h/2,init_x+addx,init_y+addy)##先畫線
                
                cv.create_text(init_x+addx-10,init_y+addy-oval_h/2-10,text=sheet.cell(row=current_row,column=k).value,font=("標楷體",10))  ##上文字(標籤)
                
                cv.create_oval(init_x+addx-oval_w/2,init_y+addy-oval_h/2,init_x+addx+oval_w/2,init_y+addy+oval_h/2,fill="white")   ##畫橢圓
                
                cv.create_text(init_x+addx,init_y+addy,text=sheet.cell(row=current_row-1,column=k).value,font=("標楷體",10))   ##上文字(詞彙)

                items+=1
                
                print(sheet.cell(row=current_row-1,column=k).value,"，座標:",str(init_x+addx),str(init_y+addy))
                
        if max_x<= init_x+addx:   ##如果x最大值比起使值+增加幅度還要小，則x最大值會增加(讓canvas邊界寬度增加)
            max_x= init_x+addx+50
         
        
      
        ##重設座標增加幅度
        addx = 0
        addy = 120
        ##重設座標增加幅度

        count+=1  ##下一句話
         
        if items>=7: ##如果items>=7，則下一句話的起始，是從起始的x軸的下面開始畫
            
            
            init_x = 300
            init_y = next_y+50
        
        if(count%2==1):       #如果count除以2餘數為1，則下一句話的起始，是從起始的x軸的下面開始畫
            init_x = 300
            init_y = next_y+50
            
            
        else:
            if not items>=7:
                init_x = next_x+550   #如果count除以2餘數不為1，則往旁邊畫圖
            

        cv.configure(height=next_y+oval_h,width=max_x)   #canvas重新調整大小
    
            
        print()
    


    
    
    
