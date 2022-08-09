## 用途: 讀取excel做處理。
## 處理完後，輸出list給p3_output和p3_show2和p3_show3

##read(keys1,keys2,keys3,keys4,keys5,sheet,output_type): 讀取excel，並做以下處理
	

## load_excel(i,sheet,is_first) :
## 把excel讀入，根據標籤不同，存到不同的list

## compare_check() 
## 判斷這句話如果沒有主事者接受者，但是有兩個時間點或物品，且是做比較，則這句話會把兩個時間點變成主事者接受者
## 例如:小明早上看了3本書，下午比早上多看2本書。

## chenge_plus(keys4,is_first)
## 如果有"比"，且第一句話的運算符號不為=，要改成=

## check_plus(output_type,is_first)
## 處理運算符號(+-=大小...)，依照變數去做改變

## name_check(is_first,keys4) 
## 處理主事者接受者。主要處理時機是: 存放主事者接受者的list裡面數量為0

## move_list2()
## 如果這句話有物品，則暫存這個物品
## 如果這句話沒有物品，前面有暫存物品，則把暫存的物品存入list

## check_tempname()
## 暫存主事者

## p3_move.move_adj(list2,list3)
## 把特點移動到數量的前面

## p3_move.move_item(list2,list3)
## 把物品移動到數量的前面
## 例如 : 我 有 1 顆 蘋果
## => 我 有 蘋果 1 顆

## p3_divide.devide(init_type,current_type,list1,list2,list3,list4,list5,keys1,keys2,keys3,keys4,keys5,temp_plus,is_first,temp_item,conj)
##如果這句話的單位有2個以上，會拆成好幾句
## 例如 : 我有1支原子筆和1支鉛筆 => 我有1支原子筆、1支鉛筆


#################################影響計算的參數#####################################################
# is_reverse = False  ##是否倒推計算
# is_minus = False  ##是否用原本數字下去扣
# is_plus = True   ##是否用一般計算，如果為False，且本身是+-，則要往下去找出=的項目做加減
# is_differ = False  ##是否用兩個數字相減
# is_total = False  ##判斷題目中是否有"一共"這類型詞彙，如果有，做倒推
# is_remain = False  ##判斷題目中是否有"剩下"這類型詞彙，如果有，做倒推
# change_plus = False   ##判斷題目中是否有"其中"這類型詞彙，如果有，做倒推
# plus_one = False  ##判斷是否要加1(當題目是排隊問題，問總排隊人數)
###############################影響計算的參數#####################################################


from tkinter import *
from random import randint
import pyautogui
from openpyxl import load_workbook
from . import p3_divide
from . import p3_move

############ 變數 ################################################

pronoun = ["自己","他","她","它","牠"] ##代名詞

temp_name = "" ##暫存前一句的人名
temp_plus = "" ##暫存前一句的運算符號
is_people=0 ##0:沒有主事者，1:有主事者

temp_item=""  ##暫存前一句的物品
temp_unit=""  ##暫存前一句的單位

init_type = current_type="物" ##第一句話的初始型態

is_do = False  ##判斷句子裡是否有"做"這個詞
compare= False ##判斷句子裡是否有"比"這個詞
yet = False ##判斷句子裡是否有"已經"這個詞
is_plus = True ##判斷是否為比較題，是否為一般運算
buy = False ##有"買" 這個詞
is_total = False  ##判斷句子中是否有"共有"
is_and = False

add_back = False ##判斷句子中是否有"不夠"，解題時加回去

conj = False ##是否有連接詞(目前沒用到)
change_plus=False ##是否有"其中"或"其他"，解題時用倒扣的

list1=[]   ##主事者或接受者名稱
list2=[]   ##物品 數量  單位  (名稱)
list3=[]   ##物品 數量  單位  (標籤)
list4=[]   ##+-=
list5 = []   ##動詞
list6=[]   ##特點
list7=[] ##發生地
list8 = []##時間點

temp_time = []  ##暫存時間

first_place = False  ##如果第一句沒有主事者接受者，只有發生地，將發生地轉為主事者

is_remain = False  ##剩下

pre = False  ##判斷是否有'前'這個詞性
post  = False##判斷是否有'後'這個詞性

total_list = ["共有","總共","共","一共","一共有"] ##共有

plus_one = False ##是否要加1


############ 變數 ################################################




def load_excel(i,sheet,is_first):   ##匯入詞彙進行處理

    global temp_name,temp_plus,is_people,init_type,current_type,temp_item,temp_unit

    global is_do ,yet ,compare,is_plus,temp_time,is_remain,is_total,conj,change_plus,plus_one,pre,post,first_place,buy,is_and,add_back

    global keys1,keys2,keys3,keys4,keys5

    for j in range(1,100):  ##讀excel裡面的資訊

        count = 0

        
        #is_merge=False  ##是否可合併特點和物品(沒用到)
        
        if (type(sheet.cell(row=i,column=j).value) !=str) and (type(sheet.cell(row=i,column=1).value) !=int) :  ##如果資料為空的，就退出
            
            
            break
        else:
            

            try:
                if sheet.cell(row=i,column=j).value=="特點" and  sheet.cell(row=i,column=j).value=="物品" :
                    is_merge=True
            except:
                continue

            if sheet.cell(row=i,column=j).value=="連接詞":
                print("連接詞",sheet.cell(row=i-1,column=j).value)
                #conj = True  ##沒用到
            
            
            if sheet.cell(row=i,column=j).value=="主事者" or sheet.cell(row=i,column=j).value=="接受者" :    ##取得主事者接受者，存到list1中
                
                
                if first_place==True and temp_name!="":  
                    
                    temp_name = ''

                if "的" in sheet.cell(row=i-1,column=j).value:  ##爸爸的=> 爸爸

                    t = sheet.cell(row=i-1,column=j).value

                    t = t.replace("的","")

                    list1.append(t)
                else:
                
                    list1.append(sheet.cell(row=i-1,column=j).value)

                if len(list1)==1 and sheet.cell(row=i,column=j).value=="接受者" and temp_name!="":  ##如果標籤為接受者，要在前面補上主事者(用暫存主事者)
                    list1.insert(0,temp_name)
                
                
                if sheet.cell(row=i-1,column=j).value not in  pronoun:  
                    is_people = 1   ##表示有主事者
                    
                current_type="人"  ##如果有找到主事者或接受者，這句話的類別為人

                #if is_first==True :
                  #  temp_name=(sheet.cell(row=i-1,column=j).value)

            
            elif sheet.cell(row=i,column=j).value=="特點" and "的" in sheet.cell(row=i-1,column=j).value and len(list1)==0 and len(list2)==0:
                
                
               
                #if sheet.cell(row=i,column=j).value=="特點" and "的" in sheet.cell(row=i-1,column=j).value and len(list1)==0 and len(list2)==0:


                if sheet.cell(row=i,column=j+1).value!="主事者": ##如果特點為XX的，且前面沒有存主事者或接受者或物品等，則把這個詞去掉"的"，變成主事者
                
                    n = sheet.cell(row=i-1,column=j).value
                    

                    n = n.replace("的","")
                    list1.append(n)
                
                else:
                    list2.append(sheet.cell(row=i-1,column=j).value)
                    list3.append(sheet.cell(row=i,column=j).value)
                

            elif sheet.cell(row=i,column=j).value in ["物品","數量","單位","特點","前","後"] :  
                
                ##如果是標籤在["物品","數量","單位","特點","前","後"]，則會把這個詞存到list2，標籤存到list3

                if sheet.cell(row=i,column=j).value=="前":
                    pre = True
                elif sheet.cell(row=i,column=j).value=="後":
                    post = True

                
               

                if pre==post==True:   
                    
                    ##如果這句話有"前"跟"後"，則plus_one = True，表示解題時要+1 。
                    ##例如 : 我前面排了5人，後面排了3人。總共排了多少人?

                    plus_one = True
                
                if is_remain == True and  sheet.cell(row=i,column=j).value=="特點" and sheet.cell(row=i-1,column=j).value=="剛好":
                    
                    
                    ##如果is_remain==True，表示這個題目可能是用倒扣。例如:我有一些錢，花了30元，剩下20元。我原本有幾元?
                    ##如果句子中還有剛好，則不做到扣。例如:我有100元，你給我30元，我花掉10元，剩下的錢剛好可以買1本書。這本書多少元?
                    
                    
                    is_remain = False  
                
                
                if sheet.cell(row=i,column=j).value=="單位" and sheet.cell(row=i-1,column=j).value=="個" and temp_unit=="人":
                    list2.append("人")   ##如果這句話單位是"個"，但是暫存的單位是"人"，則將"個"轉換成"人"
                    
                elif sheet.cell(row=i,column=j).value=="單位" and sheet.cell(row=i-1,column=j).value=="年" and temp_unit=="歲":
                    list2.append("歲")  ##如果這句話單位是"年"，但是暫存的單位是"歲"，則將"年"轉換成"歲"
                            
                            
                else:

                    if sheet.cell(row=i,column=j).value=="單位" and sheet.cell(row=i-1,column=j).value=="人" and temp_item=="人" and temp_unit=="個":
                         list2.append(temp_unit)  ##如果前面一句話的物品是"人"，單位是"個"，這句話單位為"人"，則把這句話單位改成"個"
                    else:
                        list2.append(sheet.cell(row=i-1,column=j).value)   ##如果標籤為物品數量單位特點...，則標籤名稱貼入list2中
                list3.append(sheet.cell(row=i,column=j).value)      ##如果標籤為物品數量單位，則標籤種類貼入list3中

    
                
            elif sheet.cell(row=i,column=j).value in ["+","-","=","大","小"]:
                
                if sheet.cell(row=i-1,column=j).value=="不夠":
                    list4.append("+")   ##如果標籤為+=，則標籤種類貼入list4中
                    list5.append(sheet.cell(row=i-1,column=j).value)  ##+-=名稱
                    temp_plus = "+"  ##暫存+-=

                elif sheet.cell(row=i-1,column=j).value in ['買','合買','買了'] and sheet.cell(row=i,column=j).value=="+" :#and temp_plus!="" and is_first==False :#(temp_unit=="元" or temp_item=="錢"):
                    ##我有X元，買了1個10元橡皮擦，剩下幾元
                    list4.append("-")
                    list5.append(sheet.cell(row=i,column=j).value)
                    temp_plus = "-"
                    buy = True  ##
##                elif sheet.cell(row=i-1,column=j).value in ['拿走'] and sheet.cell(row=i,column=j).value=="+" and list1==[]:
##                    list4.append("-")
##                    list5.append(sheet.cell(row=i,column=j).value)
##                    temp_plus = "-"
##                elif sheet.cell(row=i-1,column=j).value in ['拿出'] and sheet.cell(row=i,column=j).value=="-":
##                    list4.append("+")
##                    list5.append(sheet.cell(row=i,column=j).value)
##                    temp_plus = "+"

                elif sheet.cell(row=i-1,column=j).value in ['給','分給'] and sheet.cell(row=i,column=j).value=="-" :
                    list4.append("+")
                    list5.append(sheet.cell(row=i,column=j).value)
                    temp_plus = "+"
                
                elif sheet.cell(row=i-1,column=j).value in ['煮','煮了'] and sheet.cell(row=i,column=j).value=="+" :
                    list4.append("-")
                    list5.append(sheet.cell(row=i,column=j).value)
                    temp_plus = "-"
                
                elif sheet.cell(row=i-1,column=j).value in ['吃','吃了'] and sheet.cell(row=i,column=j).value=="+" :
                    list4.append("-")
                    list5.append(sheet.cell(row=i,column=j).value)
                    temp_plus = "-"
                
                elif sheet.cell(row=i-1,column=j).value in ['放進'] and sheet.cell(row=i,column=j).value=="+" :
                    list4.append("-")
                    list5.append(sheet.cell(row=i,column=j).value)
                    temp_plus = "-"

                elif sheet.cell(row=i-1,column=j).value in ['跳走'] and sheet.cell(row=i,column=j).value=="-" :
                    list4.append("+")
                    list5.append(sheet.cell(row=i,column=j).value)
                    temp_plus = "+"
                    
                elif sheet.cell(row=i-1,column=j).value!="比":

                    
                    
                    if buy==True and sheet.cell(row=i-1,column=j).value=="剩下" and temp_unit in list2:  ##買了X，剩下Y元。
                        ##ex : 小明有一些錢，買了1本50元的筆記本，剩下20元。小明原本有幾元?
                        ## abs(-20-50) = 70
                        
                        list4.append("-")
                        list5.append(sheet.cell(row=i-1,column=j).value)  ##+-=名稱

                        
                        
                    else:
                        
                        list4.append(sheet.cell(row=i,column=j).value)   ##如果標籤為+=，則標籤種類貼入list4中
                        
                        list5.append(sheet.cell(row=i-1,column=j).value)  ##+-=名稱
                        temp_plus = sheet.cell(row=i,column=j).value  ##暫存+-=

                    

                    if "-" in list4 and change_plus==True:
                        change_plus = False
  
                if sheet.cell(row=i-1,column=j).value in ["用" , "用掉"]:
                    
                    is_do = True
                if sheet.cell(row=i-1,column=j).value in ["剩下","剩","找回","還有"] :  ##如果是"剩下"
                    if sheet.cell(row=i-1,column=j+1).value !="的":
                        is_remain = True

                if sheet.cell(row=i-1,column=j).value=="比":

                    
                    compare = True
                    is_plus = False

                if sheet.cell(row=i-1,column=j).value in total_list:# and is_first==True:
                    
                    is_total = True  ##之後解題會用到(總和=>用扣的)

                if sheet.cell(row=i-1,column=j).value=="和":

                    is_and = True

                if sheet.cell(row=i-1,column=j).value in ["不夠"]:
                    
                    add_back = True
                    

##                  
            elif sheet.cell(row=i,column=j).value=="發生地":
                list7.append(sheet.cell(row=i-1,column=j).value)

                
            elif sheet.cell(row=i,column=j).value=="虛詞":
                
                if sheet.cell(row=i-1,column=j).value=="做":
                    is_do = True
                elif sheet.cell(row=i-1,column=j).value=="其中" or sheet.cell(row=i-1,column=j).value=="其他":
                        
                    if "-" not in list4:
                        change_plus=True  ##之後的+-會做對調
                elif sheet.cell(row=i-1,column=j).value in total_list:
                    
                    is_total = True  ##之後解題會用到(總和=>用扣的)
                    

            elif sheet.cell(row=i,column=j).value=="時間點":
                
                if sheet.cell(row=i-1,column=j).value=="已經":
                    yet = True
                list8.append(sheet.cell(row=i-1,column=j).value)
    
    
    
    
    
    
    return list1,list2,list3,list4,list5,list6,list7,list8
    

def reset():  ##重置

    global temp_name,temp_plus,is_people,init_type,current_type,temp_item,temp_unit,is_total

    global is_do ,yet ,compare,is_plus,temp_time,is_remain,conj,is_and

    global list1,list2,list3,list4,list5,list6,list7,list8

    temp_name = "" ##暫存前一句的人名
    temp_plus = ""
    is_people=0
    temp_item=""
    temp_unit=""

    init_type = current_type="物"

    is_do = yet =compare= False  ##做、比較
    is_plus = True

    is_total = False  ##判斷句子中是否有"共有"

    is_and = False

    conj = False

    list1=[]   ##主事者或接受者名稱
    list2=[]   ##物品 數量  單位  (名稱)
    list3=[]   ##物品 數量  單位  (標籤)
    list4=[]   ##+-=
    list5 = []   ##動詞
    list6=[]   ##特點
    list7=[] ##發生地
    list8 = []##時間點

    temp_time = []  ##暫存時間
    
    first_place = False

    is_remain = False  ##剩下
    
    
    
def chenge_plus(keys4,is_first):  ## 如果有"比"，且第一句話的運算符號不為=，要改成=
    

    if len(keys4)>0 and is_first==False:  ##如果Keys4數量>0，且非第一句，則要判斷裡面是否有=
        
         list_ = keys4[0]
         
         if '=' not in list_:
             
             list_[0] = '='
             
             return
             
                
    
    return

def change_and(keys4):

    if len(keys4)>0:

        list_ =  keys4[0]
        
        if '=' not in list_:

            list_[0] = '='

            return


    return
    

def read(keys1,keys2,keys3,keys4,keys5,sheet,output_type):

    temp1=[]
    temp2=[]
    temp3=[]
    temp4=[]
    temp5=[]


    if output_type!="":   ##當輸出格式改變時，重置變數
        reset()

    global temp_name,temp_plus,is_people,init_type,current_type,temp_item,temp_unit,is_total

    global is_do ,yet ,compare,is_plus,temp_time,is_remain,conj,first_place,is_and,add_back

    global list1,list2,list3,list4,list5,list6,list7,list8

    
    
    init_type = "物"  ##主事者原本的詞性(預設:人)
    current_type = "物"  ##主事者的詞性(預設:人)

    temp_item=""   ##暫存物品名稱
    temp_unit = ""

    is_first=True   ##判斷是否為第一句
    is_do = False   ##判斷是否有'用掉'或'用'
    

    temp_time = []##暫存時間
    
    
    for i in range(4,100,2):   ##最多7句話

        

        list1=[]   ##主事者或接受者名稱
        list2=[]   ##物品 數量  單位 特點 (名稱)
        list3=[]   ##物品 數量  單位 特點 (標籤)
        list4=[]   ##+-=
        list5 = []   ##動詞
        list6=[]   ##特點
        list7=[] ##發生地
        list8 = []##時間點

##        is_do = False   
##        yet = False  ##如果句子裡面有"已經"
##        is_plus = True    ##如果句子中有"比"，則is_plus=False
##        compare =  False   ##是否是比較
        
        

        if (type(sheet.cell(row=i,column=1).value) !=str) and (type(sheet.cell(row=i,column=1).value) !=int) :  ##如果資料為空的，就退出
            break

            
        load_excel(i,sheet,is_first)  ##讀取excel的詞彙與詞性

        

        check = check_enter(list1,list2,list3,list4,list5,list6,list7,is_first,temp_name,compare,is_and)  ##判斷是否做後續處理


        
        if check==False:
            
            if len(list7)!=0 and temp_name=='' :   ##如果有發生地，且暫存主事者為空，把暫存的主事者指定為這個發生地
                
                temp_name=list7[0]
                check_name=True  ##沒用到
                first_place = True

                
        
        if check==True:
            
           
            if compare ==True:  ##如果句子中有"比"這個詞

                
                compare_check()   ##處理兩個主事者(物品或時間點)的狀況
                
                chenge_plus(keys4,is_first)  ##如果有"比"，且第一句話的運算符號不為=，要改成=

            

                

                

            check_plus(output_type,is_first) ##處理運算符號
                
            
            name_check(is_first,keys4)  ##處理主事者接受者
            
            
            move_list2()  ##把物品數量單位特點等，做處理，例如暫存物品等
            
            check_tempname()  ##暫存主事者名稱
                
           
            list2,list3 = p3_move.move_adj(list2,list3)  ##移動特點,時間點,發生地到前面
            
            list2,list3 = p3_move.move_item(list2,list3)  ##移動物品到前面

            if len(list7)!=0:
                for x in range(len(list7)-1,-1,-1):

                    list2.insert(0,list7[x])
                    list3.insert(0,"發生地")
            
            
            if len(list1)==1:  ##主事者接受者數量為1

                
                
                
                if len(list4)==0:
                    list4.append("=")
                    
                
                
                if list3.count("單位")>1 :
                    
                    
                    keys1,keys2,keys3,keys4,keys5=p3_divide.devide(init_type,current_type,list1,list2,list3,list4,list5,keys1,keys2,keys3,keys4,keys5,temp_plus,is_first,temp_item,conj)
                    if is_and == True:

                        change_and(keys4)
                    
                else:
                    
                    keys1.append(list1)   
                    keys2.append(list2)
                    keys3.append(list3)
                    keys4.append(list4)
                    keys5.append(list5)

                      

                    
            
            else:  ##主事者接受者數量為2

               
                    
                
                temp1.append(list1) 
                temp2.append(list2)
                temp3.append(list3)
                temp4.append(list4)
                temp5.append(list5)

            init_type = current_type

   
       
        elif len(list1)==2:
            keys1.append(list1)   
            keys2.append(list2)
            keys3.append(list3)
            keys4.append(list4)
            keys5.append(list5)


        is_first=False  ##是否為第一句(第一句跑完後為false)
        
            
    if len(temp1)!=0:

                
            
        for i in range(len(temp1)):

                
            keys1.append(temp1[i])   
            keys2.append(temp2[i])
            keys3.append(temp3[i])
            keys4.append(temp4[i])
            keys5.append(temp5[i])

            #把所有小的list存到大的list中
                
            
            
            ###把標籤為物品的放到數量前面

    
   
    return keys1,keys2,keys3,keys4,keys5,is_plus,is_remain,is_do,is_total,change_plus,plus_one,add_back



def move_list2():  ##暫存物品、加入物品

    global temp_name,temp_plus,is_people,init_type,current_type,temp_item,temp_unit

    global is_do ,yet ,compare,is_plus,temp_time,is_and

    global list1,list2,list3,list4,list5,list6,list7,list8

   

    if list3.count("物品")!=0:# and list3.count("單位")!=0:   ##如果前一句話有物品，則暫存此物品
        
        for x in range(len(list3)):
            
            
            if list3[x]=="物品":
                
                temp_item=(list2[x])

            
                
            for x in range(len(list3)):
                if list3[x]=="單位":
                    temp_unit=(list2[x])
                    if temp_unit=="元":
                        temp_item=""

    else:
        ##暫存單位
        for x in range(len(list3)):
            if list3[x]=="單位":
                temp_unit=(list2[x])
                
            
    
    if list3.count("物品")==0 and temp_item!="" and list3.count("單位")!=0 :   ##如果只有單位，再加入物品

        
        
        a = list3.index("單位")
        a = list2[a]

       
        
        if a==temp_unit or temp_unit=="":
            
            list2.append(temp_item)
            list3.append("物品")

    

    if len(list8)!=0:
        for x in list8:
            if x not in temp_time:
                temp_time.append(x)
            list2.insert(0,x)
            list3.insert(0,"時間點")

    

def check_tempname():

    global temp_name,temp_plus,is_people,init_type,current_type,temp_item,temp_unit

    global is_do ,yet ,compare,is_plus,temp_time,is_and

    global list1,list2,list3,list4,list5,list6,list7,list8

    if len(list1)==1 and list1[0] not in pronoun:   ##暫存名稱，如果這個詞本身是代名詞，則不會存入

                
        temp_name = list1[0]

        
  
    elif temp_name=="" and list1[0] in pronoun :
        temp_name = list1[0]


    

def check_enter(list1,list2,list3,list4,list5,list6,list7,is_first,temp_name,compare,is_and):  ##判斷是否這句話要繼續做處理

    
    check = False
    
    
    
    if  ((len(list1)>0 or temp_name!="") and (list3.count("物品")!=0 or list3.count("單位")!=0)):  ##主事者數量>0或temp_name不為空，且物品或單位不為空
        
        check = True
        return check

    elif (list3.count("物品")==2 and compare==True):  ##如果是兩物品比較
        check = True
        return check

    elif (len(list1)==2):   ##如果是兩主事者接受者
        check = True
        return check

    elif (list3.count("物品")!=0 and list3.count("數量")!=0):  ##如果物品跟數量不為空

##        if temp_name=="" and is_first==True and len(list4)==0:
##            check = False
##            return check
        
        check = True
        return check
    elif len(list1)!=0 and is_first==True:   ##如果第一句有主事者
        check = True
        return check

    elif (list3.count("物品")!=0 and list3.count("單位")!=0):  ##如果物品單位不為空

        check = True
        return check

    elif (len(list7)!=0 and (list3.count("物品")!=0 or list3.count("單位")!=0)):##如果有發生地，且物品或單位不為空

        check = True
        return check
    elif list3.count('物品')==2 and is_first==True:
        check = True
        return check
    elif "數量" in list3:
        check = True
        return check
    else:
        return check

def compare_check(): ##處理主事者接受者，(前面所存的主事者接受者list數量為0時，且有"比"這個詞)

    global temp_name,temp_plus,is_people,init_type,current_type,temp_item,temp_unit

    global is_do ,yet ,compare,is_plus,temp_time,is_and

    global list1,list2,list3,list4,list5,list6,list7,list8

    
    ##我早上買了5顆蘋果，下午比早上多買2顆
    
    if len(list1)==0 and len(list8)==2:##人或時間點
        
        for x in list8:
            
            #if x not in temp_time:
                
            list1.append(x)

            a = list8.index(x)

            #del list8[a]
        list8 = []

    elif len(list1)==0 and list3.count("物品")==2:  ##主事者接受者數量為0，物品數量為2

        
        
        for i in range(2):

            a = list3.index("物品")
            
            list1.append(list2[a])
            del list2[a]
            del list3[a]

            

            if a>1 :

                if list3[a-1]=="單位" and int(list2[a-2])==1:
                    del list2[a-1]
                    del list3[a-1]
                    del list2[a-2]
                    del list3[a-2]


    elif len(list1)==0 and list3.count("物品")==1:
            if temp_name!=0:
                list1.append(temp_name)
                
            

            a = list3.index("物品")
                
            list1.append(list2[a])
            del list2[a]
            del list3[a]
        
    elif list3.count("物品")>2:

        it= []

        for i in range(len(list3)):

            if list3[i]=='物品':

                it.append(list2[i])

        

        for i in range(len(list3)):

            if list2[i]==temp_name:

                
                list1.append(temp_name)
                del list3[i]
                del list2[i]
                break

        for i in range(len(list3)):

            if list3[i]=='物品':
                
                if temp_name in it and it.index(temp_name)!=0:
                    list1.insert(0,list2[i])
                else:
                    list1.append(list2[i])
                del list3[i]
                del list2[i]
                break
    

def name_check(is_first,keys4):  ##處理主事者接受者(這句話主事者接受者數量為0)

    global temp_name,temp_plus,is_people,init_type,current_type,temp_item,temp_unit

    global is_do ,yet ,compare,is_plus,temp_time,first_place,is_and

    global list1,list2,list3,list4,list5,list6,list7,list8

    
    
    
    if len(list1)==0 and temp_name!="":    ###如果沒有主事者接受者，但有暫存主事者接受者
        
        
        if init_type==current_type=="物" and temp_item==""  and list3.count("物品")==1:
            
            i = list3.index("物品")
            item = list2[i]
            list1.append(item)
            del list2[i]
            del list3[i]
        else:      
            list1.append(temp_name)

            
        
        if temp_name in list2:
            a = list2.index(temp_name)
            del list2[a]
            del list3[a]
            
        
        if is_do ==True:   ##物品移動到主事者
            
            if list3.count("物品")==1:
                a = list3.index("物品")
                    
                list1.append(list2[a])
                del list2[a]
                del list3[a]

    elif len(list1)==0 and len(list7)!=0 :  ##如果發生地不為空，則當作主事者
        
        
        list1.append(list7[0])
        temp_name = list7[0]
        current_type="人"
        del list7[0]
        
        #first_place = True
        
        #print(first_place)
        

    elif  is_people==0 and temp_name=="" and len(list1)==0:  ##如果前面一句沒有主事者接受者

        
                    
        current_type = "物"

        
        if (list4[0]=="=" and compare==False) or (is_first==True):  ##如果+-=為=，則第一個物品當作主事者

            
            for k in list3:   ##如果list3裡面有"物品"，則把list2相對應的標籤放到list1(主事者接受者)第一項中
                if k=="物品":

                    
                    list1.insert(0,list2[list3.index(k)])
                    temp_name = list2[list3.index(k)]
                    
                    del list2[list3.index(k)]
                    del list3[list3.index(k)]
                    
                    break

        else:
            
            
            for k in list3:   ##如果list3裡面有"物品"，則把list2相對應的標籤放到list1(主事者接受者)第一項中
                if k=="物品":
                    if list2[list3.index(k)] not in list1:
                        list1.insert(0,list2[list3.index(k)])

                        del list2[list3.index(k)]
                        del list3[list3.index(k)]
                    
            #for k in range(len(list3)-1,-1,-1):
                #if list3[k]=="物品":
                    #del list3[k]

                    
                    
            
            if len(list1)==2:   ##如果主事者接受者數量為2，則兩個元素互換(原本的標籤因為都是insert到第一項，所以第一項跟第二項是相反的)
                
                temp = list1[0]
                list1[0]=list1[1]
                list1[1]=temp

        
        
            
    
            
                
##    if len(list7)!=0 and len(list1)!=0:  ##如果list1裡面不為空，則list7(發生地)存到list6(特點)
##        for x in list7:
##            list6.append(x)

    if len(list1)==1 and  len(list4)>0 and list4[0] in ["+","-"] and temp_name!=list1[0] and list1[0] not in pronoun and temp_name!="" :   ##如果主事者數量為1，包含+-符號，則加入暫存名稱(表示加減性)
        
        if current_type=="物" == init_type:
            list1.insert(0,temp_name)
        
        else:
            
            if len(keys4)>0:
                list_ = keys4[0]
                if list_[0]!=list4[0]:
                    list1.append(temp_name)

        
    
    
    pronouns = 0

    for x in list1:

        if x in pronoun:

            pronouns+=1
            
    if pronouns<2:
        for x in range(len(list1)):   ##判斷代名詞，如果代名詞有在pronoun這個list中，則轉換成主事者接受者名稱
            
            

            if list1[x] in pronoun :


                list1[x] = temp_name
    
    
    
    if len(list1)==1 and len(list4)>0 and list4[0] in ["+","-"] :
        
        if init_type== "物" and current_type!= init_type and temp_name!=list1[0] and temp_name!="":
            
            list1.append(temp_name)
    
    if (len(list4)==2 and list4[0]=="+" and list4[1]=="-" and "物品" in list3) and is_first==False :  ##把物品移動到主事者或接受者
        
        eq = False
        for i in keys4:
            
            if '=' in i:  ##確認不是第一句話，且=有在list中
                eq = True
                
        if eq==False:
            return
        del list4[0]
        
        a = list3.index("物品")
        list1.append(list2[a])
        
        del list2[a]
        del list3[a]
        
    elif len(list1)==2 and  len(list4)==2 and list4[0]=="+" and list4[1]=="-":
        
        del list4[0]

    

def check_plus(output_type,is_first):  ##處理+-=

    
    global temp_name,temp_plus,is_people,init_type,current_type,temp_item,temp_unit

    global is_do ,yet ,compare,is_plus,temp_time,change_plus,is_and   ##output_type輸出類型(圖或是txt)

    global list1,list2,list3,list4,list5,list6,list7,list8

    
    
    if len(list4)==0 and list3.count("單位")>1:   ##增加+-=
        if temp_plus!="":
           list4.append(temp_plus) 
        else:
            list4.append("=")
            temp_plus = "="

        
        
        
    elif len(list4)==1 and yet ==True and is_first==False:  ##已經
        
        if list4[0] in ["=","+"]  :
            list4[0]= "-"

    elif len(list4)==1 and is_plus==False:   ##如果句子中有比較，則+-互換，運算才正確
      
        if list4[0]=="-":
            list4[0]="+"
        elif list4[0]=="+":
            list4[0]="-"

##    elif len(list4)==1 and change_plus==True:
##
##        list4[0]="-"

    if len(list1)==2 and len(list4)==0:
        if temp_plus!="":
            
            list4.append(temp_plus)
        else:
            list4.append("=")

    if len(list1)==2 and list4[0]=="=" and is_first==True:   ##如果第一句話，主事者接受者有兩個，且符號為=，則可能為減法
        
        change_plus=True
        
        
    if temp_name=="" and len(list1)==2:   ##如果暫存名稱為空，且主事者接受者有兩人，則先暫存第一個人
        temp_name=list1[0]

    if len(list4)==2 and "有" in list5:
        a = list5.index("有")
        del list4[a]
        del list5[a]
    elif len(list4)==2 and ("剛好可以" in list5 or "剛好" in list5 or "正好" in list5 or "恰好" in list5):
            ##如果list5元素有兩個(表示list4元素也有兩個)，且其中包含["剛好可以","剛好","正好","恰好"]中的任一個，則會替除掉不包含的詞
        for x in range(len(list5)):
            if list5[x] not in ["剛好可以","剛好","正好","恰好"] :
                
                del list4[x]
                del list5[x]

            else:
                temp_plus=list4[x]  ##如果list5元素在["剛好可以","剛好","正好","恰好"]，暫存單位

   
        
    
    if "數量" in list3 and output_type=="txt":
        
        
        if "大" in list4 or "小" in list4:

            

            for x in range(len(list4)):  ##輸出給project4，大:+，小:-

                
                if list4[x] in ["大","小"]:

                   

                    if list4[x]=="大":
                        
                        list4[x]="-"
                        
                    elif list4[x]=="小":
                        
                        list4[x]="+"
            

            for x in range(len(list4)-1,-1,-1):  ##比較的句子，把不相干的符號拿掉

                if list4[x] not in ["+","-"]:
                    
                    del list4[x]
                    del list5[x]

    
    
    if len(list4)==0:
        if temp_plus!="":
            list4.append(temp_plus)
        else:
            list4.append("=")
            temp_plus="="
