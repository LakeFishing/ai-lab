from tkinter import *
from random import randint
import pyautogui
from openpyxl import load_workbook


if __name__ == '__main__':
    wb = load_workbook(filename = 'proj2.xlsx')
    sheet = wb.active

    unit = ""  ##單位
    units = []  ##存放單位(至多兩個) (目前沒用到)

def find_unit(list2,list3):

    ##先找出解題敘述中的單位，之後處理單位會用到
    
    global unit,units
    
    
    unit = ""
    units = []
    
    for i in range(1,100):   ## 

        

        if (type(sheet.cell(row=1,column=i).value) !=str) and (type(sheet.cell(row=1,column=i).value) !=int) :  ##如果資料為空的，就退出
                break
        else:

            if sheet.cell(row=2,column=i).value=="單位":
                unit = sheet.cell(row=1,column=i).value

                if unit=="錢":
                    unit = "元"
                    
                if sheet.cell(row=2,column=i).value in ["大","小"]:
                    unit = ""

                if unit!="":
                    units.append(unit)

    


    if unit=="" or unit=="種" or unit not in list2:  ##如果問題沒有單位，則從題目中找單位

        for i in range(len(list3)):

           if list3[i]=="單位":

               unit = list2[i]
    
    

def devide(init_type,current_type,list1,list2,list3,list4,list5,keys1,keys2,keys3,keys4,keys5,temp_plus,is_first,temp_item,conj):  ##p3_read3拆解句子用

    ##根據單位分割成不同的部分

    find_unit(list2,list3)  ##先處理單位，找出最多兩個單位，並存list中

    
    pos = 0  
    time = 0

    t1 = [] ##暫存主事者
    t2 = []  ##暫存物品數量單位...
    t3 = []  ##暫存物品數量單位...(詞性)
    t4 = []  ##暫存+-=

    items = []  ##暫存物品

    change_plus = False
    
    
    

    for i in range(len(list3)):    ##暫存物品，之後要挪動位置用
        if list3[i]=="物品":
            items.append(list2[i])
            
    

#########################   先把list4中的元素數量，調整到和list3中"單位"數量一樣   ###########################################################################
   



    if len(list1)==1:   ##如果主事者數量為1，則進行下面判斷

        if list3.count("單位")<1:  ##如果沒有運算符號，則要先補一個進去

            for i in range(len(list3)):   ##判斷是否要換+- : 如果list2的單位和解題的單位不同，且list4不為["+","+"]
                
                if list2[i]!=unit and list3[i]=="單位" and list4!=["+","+"]:
                    change_plus = True

 
        if list3.count("單位")>=1:  ##如果"單位"的數量>=1

            if len(list4)==0:   ##如果list4沒有+-=

                if temp_plus!="":  ##如果temp_plus不為空，表示有暫存+-=

                    list4.append(temp_plus)  #list4存入+-=
                else:
                     list4.append("=")   ##如果沒有，則存入=
                
            
            if len(list4)!=0:   ##如果list4有+-=

                if len(list4)== list3.count("單位"):
                    if temp_plus!="":
                            
                        if list4[0]=="+" and change_plus==True and is_first==False and temp_item=="":   ##如果不是第一句，且裡面有不同單位，lis4包含+，則要換成-
                                
                            list4[0]="-"
                            temp_plus="-"


                            
                if len(list4)<list3.count("單位"):  ##如果運算符號數量小於list3中"單位"的數量，則list4要補運算符號進去，直到跟list3中"單位"的數量一樣多
                
                
                    for i in range(list3.count("單位")):   ##根據"單位"的數量，在list4存+-=，直到list4裡面元素數量跟list3裡面"單位"數量一樣多
                        
                        if len(list4)<list3.count("單位"):

                           
                            
                            if temp_plus!="":
                               

                                if list4[0]=="+" and change_plus==True and is_first==False:   ##如果不是第一句，且裡面有不同單位，lis4包含+，則要換成-
                                    
                                    list4[0]="-"
                                    temp_plus="-"
                                    
                                list4.append(temp_plus)

                            else:
                                list4.append("=")

                elif len(list4)>list3.count("單位"):     ##運算符號比"單位"數量多，則要刪減到一樣

                    for i in range(len(list4)-1,-1,-1):

                        del list4[i]

                        if len(list4)==list3.count("單位"):

                            break
                
                for i in range(len(list2)-1,-1,-1):   ##先在單位後面加入主事者，方便之後進行斷句判斷

                    if list2[i]==unit and i<len(list2)-1:
                        
                        if conj==True and list3.count("物品")==1 and list3.count("物品")<list2.count(unit):

                            x = list3.index("物品")

                            item = list2[x]

                            del list2[x]
                            del list3[x]
                            
                            list2.insert(i+1,item)

                            list3.insert(i+1,'主事者')


                        else:
                            

                            list2.insert(i+1,list1[0])

                            list3.insert(i+1,'主事者')

                
                        
                list2.insert(0,list1[0])  ##第一個位置也要加入主事者
                list3.insert(0,'主事者')
                
                

                temp_pos = [] ##暫存跑過的位置


#########################根據單位數量，跑迴圈，把句子拆解成多句###########################################################################

                for i in range(list3.count("單位")):  

                    

                    t1 = []  ##主事者
                    t2 = []  ##物品數量單位
                    t3 = []  ##物品數量單位(詞性)
                    t4 = []  ##+-=

                    t4.append(list4[i])
                    
                    

                    for j in range(pos,len(list3)):   ##從pos開始，到list3的結尾，t2和t3存放詞彙和詞性，pos預設為0

                        

                        if j in temp_pos:
                            break
                        else:
                            temp_pos.append(j)

                        if list3[j]=="主事者":

                            t1.append(list2[j])
                        else:
                            t2.append(list2[j])
                            t3.append(list3[j])

                        

                        time+=1



                        if j<len(list3)-2:

                            if list2[j]==unit and list3[j+1]=="主事者":
                                

                                ##如果j小於list3長度-2，且list3[j+1]為數量，list3[j]在["物品","數量","單位"]中，則，則停止迴圈
                                
                                pos = time

                                
                                break


                    keys1.append(t1)  ##主事者
                    keys2.append(t2)  ##物品數量單位(名稱)
                    keys3.append(t3)  ##物品數量單位(標籤)
                    keys4.append(t4)  ##+-=
                    
                        
                 
    
                    
          
   
    return keys1,keys2,keys3,keys4,keys5














