## 把excel讀取的資料做處理，並且連結

## read3(sheet) ##讀取excel，並做處理，使用下面兩個function

## p3_read.read ##讀取excel，存到list中

##create(keys1,keys2,keys3,keys4,keys5,f,is_plus,is_remain,is_do,is_total,change_plus,plus_one):
##連結成語意網路

#################################影響計算的參數#####################################################
# is_reverse = False  ##是否倒推計算
# is_minus = False  ##是否用原本數字下去扣
# is_plus = True   ##是否用一般計算，如果為False，且本身是+-，則要往下去找出=的項目做加減
# is_differ = False  ##是否用兩個數字相減
# is_total = False  ##判斷題目中是否有"一共"這類型詞彙，如果有，做倒推
# is_remain = False  ##判斷題目中是否有"剩下"這類型詞彙，如果有，做倒推
# change_plus = False   ##判斷題目中是否有"其中"這類型詞彙，如果有，做倒推
# plus_one = False  ##判斷是否要加1(當題目是排隊問題，問總排隊人數)
# add_back = False  ##判斷題目中是否有"不夠"這類型詞彙，如果有，加回去
# no_have = False ##判斷題目中是否有"沒有、沒"這類型詞彙，有則做特殊計算
###############################影響計算的參數#####################################################


from re import T
from tkinter import *
from random import randint
import pyautogui
from openpyxl import load_workbook
from . import p3_divide

from . import p3_read
from . import p3_move


output_type = "txt"

def read3(sheet):
    
    print("p3_output")
    print()

    ##讀取excel，並做處理 : 題目&解題
    
    
    ##連結邏輯:
    ##資料結構為:list [標籤名稱，標籤位置，上一個標籤的位置，下一個標籤的位置 ]
    
    ##每一句話，先把讀出來的標籤依照不同分類，暫時存到不同的list中(list1~list5)，接者再存到(keys1~keys5)
    ##所有句子讀完後，才進行連結處理
    
    ##首先判斷主事者接受者，給定座標，把座標跟名稱存放在coor_name中，之後處理標籤會從中比對，如果有找到對應的標籤，從此標籤座標開始連結
    ##之後處理+-=標籤，給定座標，並且存在前面標籤list的第4項(下一個標籤的位置)
    
    ##接下來從物品,數量,單位的list中找出標籤來處理，給定座標，放在coor_item中。
    ##此標籤座標會存入前面標籤list的第4項(下一個標籤的位置)以及後面標籤list的第3項(前一個標籤的位置)
    
    ##如果主事者接受者無法從coor_name對應到，則從coor_item中進行比對，如果有找到對應之標籤，就進行連結
    

    ###範例:我有15個水果，你給我3個水果，我共有多少水果?
    ###標籤名稱:我,=,15,個,水果，你,給,我,3,個,水果
    ###標籤種類:主事者,=,數量,單位，主事者,-,接受者,數量,單位,物品
    
    ##第一句:先將"我"給定位置(a0)，並且存在coor_name。接者處理"="，最後是水果,15,個，並且存在coor_item中
    
    ##第二句:主事者接受者有"你"跟"我"，"我"在coor_name中有找到對應之標籤，因此產生連結，先固定其座標，接者將"你"存入coor_name
    ##最後是水果,3,個，存在coor_item中

    
    q = open('project3/question.txt','w+',encoding='utf-8')
    f = open('project3/word.txt','w+',encoding='utf-8')    ##開啟txt檔，之後要寫入資料

    f.truncate(0)
    q.truncate(0)

    word = []
    
    
    keys1 = []  ##主事者或接受者名稱

    keys2 = []  ##物品 數量  單位  (名稱)
        
    keys3 = []  ##物品 數量  單位   (標籤)

    keys4 = []   ##+-=

    keys5 = []  ##+-=名稱

    q1 = []  ##存問題之標籤名稱
    
    q2 = []  ##存問題之標籤種類

    #q_list=["主事者","+","-","=","物品","單位","總和"]

    coor_name = []   ##主事指者接受者的名稱及位置

    coor_item = []   ##物品數量單位的名稱及位置

    coor_plus = []  ##+=的位置

    output_list=[]  ##最後輸出的list

    num = ["a","b","c","d","e","f","g","h","i","j","k","l","m","n","o","p","q","r","s","t","u","v","w","x","y","z"]
    ##每一個句子的代碼(例如 第一句的第一個標籤為a0，第二句的第二個標籤為b1)

    name = ""  ##主事者接受者(只有一人)
    
    name1 = "" ##主事者接受者(有兩人)
    name2 = "" ##主事者接受者(有兩人)

    pre=[]  ##暫存前一個或前兩個標籤的位置

    naxt=[]  ##暫存下一個標籤的位置

    linked = 0 ##辨識有無連結 0:無 ; >1:有

    ##辨識有無連結 0:無 ; >1:有   (當主事者接受者數量為2時，用兩個變數去判斷)
    linked1 = 0  
    linked2 = 0
    ##辨識有無連結 0:無 ; >1:有   (當主事者接受者數量為2時，用兩個變數去判斷)

    pronoun = ["自己","他","她","它","牠"] ##代名詞
    temp_name = "" ##暫存前一句的人名
    temp_plus="="

    init_pos = "" ##這句話的第一個標籤

    is_first = True  ##是否為第一句

    is_people=0 ##是否為人

    init_type = "人"  ##主事者原本的詞性(預設:人)
    current_type = "人"  ##主事者的詞性(預設:人)
    temp_item=""   ##暫存物品
    temp_unit=""  ##暫存單位


    temp_time = []  ##暫存時間

    total_list = ["共有","總共","共","一共","一共有"] ##共有

    
    
    #isf = True  ##判斷是否是第一句(沒用到)

    for i in range(1,50):
        
        if (type(sheet.cell(row=1,column=i).value) !=str) and (type(sheet.cell(row=1,column=i).value) !=int) :  ##如果資料為空的，就退出
            break
        else:
            
            q1.append(sheet.cell(row=1,column=i).value)   ##存問題之詞彙名稱
            q2.append(sheet.cell(row=2,column=i).value)   ##存問題之標籤種類

    for i in q1:   ##寫入問題標籤(名稱)
        q.write(str(i)+" ")
    q.write("\n")

    for j in q2:    ##寫入問題標籤(種類)
        q.write(str(j)+" ")
    q.write("\n")

    

    keys1,keys2,keys3,keys4,keys5,is_plus,is_remain,is_do,is_total,change_plus,plus_one,add_back,no_have = p3_read.read(keys1,keys2,keys3,keys4,keys5,sheet,"txt")
    print("讀取excel，做初步處理")
    
    print("keys1",keys1)
    print("keys2",keys2)
    print("keys3",keys3)
    print("keys4",keys4)
    print()
    print("進行連結處理")
    create(keys1,keys2,keys3,keys4,keys5,f,is_plus,is_remain,is_do,is_total,change_plus,plus_one,add_back,no_have)  ##進行連結處理
    
        



    
def create(keys1,keys2,keys3,keys4,keys5,f,is_plus,is_remain,is_do,is_total,change_plus,plus_one,add_back,no_have):##連結成語意網路

    

    coor_name = []   ##主事指者接受者的名稱及位置

    coor_item = []   ##物品數量單位的名稱及位置

    coor_plus = []  ##+=的位置

    output_list=[]  ##最後輸出的list

    num = ["a","b","c","d","e","f","g","h","i","j","k","l","m","n","o","p","q","r","s","t","u","v","w","x","y","z"]
    ##每一個句子的代碼(例如 第一句的第一個詞為a0，第二句的第二個詞為b1)

    name = ""  ##主事者接受者(只有一人)
    
    name1 = "" ##主事者接受者(有兩人)
    name2 = "" ##主事者接受者(有兩人)

    pre=[]  ##暫存前一個或前兩個標籤的位置

    naxt=[]  ##暫存下一個標籤的位置

    linked = 0 ##辨識有無連結 0:無 ; >1:有

    ##辨識有無連結 0:無 ; >1:有   (當主事者接受者數量為2時，用兩個變數去判斷)
    linked1 = 0  
    linked2 = 0
    ##辨識有無連結 0:無 ; >1:有   (當主事者接受者數量為2時，用兩個變數去判斷)

    pronoun = ["你","我","他","她","它","牠"] ##代名詞
    temp_name = "" ##暫存前一句的人名
    temp_plus="="  ##暫存前一句的+-=

    init_pos = "" ##這句話的第一個標籤

    is_first = 0   

    

    init_type = "人"  ##主事者原本的詞性(預設:人)
    current_type = "人"  ##主事者的詞性(預設:人)
    temp_item=""  ##暫存物品
    temp_unit=""##暫存單位

     

    total_list = ["共有","總共","共","一共","一共有"] ##共有
    
   
    
    for i in range(len(keys1)):   ##每一句話去分析，並且給標籤
        print()
        print("第",str(i+1),"句")

        is_first+=1

        count = 0  ##計算第幾個標籤
        
        pre.clear()  ##清空pre(前一項位子暫存) list
        
        linked = 0  ##清空連結次數
        
        linked1 = 0  ##清空連結次數
        linked2 = 0  ##清空連結次數

        
        
        
       ##先處理主事者接受者
        print("處理主事者接受者")
        if len(keys1[i])==1:  ##主事者接受者數量為1
        
            print("主事者接受者數量為1")
            
            name = keys1[i][0]  ##暫存名稱
            
            label = num[i]+str(count)  ##暫存標籤位子(例如a1,a2,a3...)
            
            

            if is_first==1:
                init_pos = label  ##設定起始的位置
            
            
            for x in coor_name:
                if x[0]==name:     ##有找到一樣名稱的標籤
                    label = x[1]    ##暫存標籤位子
                    linked += 1    ##表示有連結
                    init_pos = label  ##設定起始的位置
                    print("產生連結，",name)
            
            
                
            if linked ==0:  ##如果沒有連結，則存入新的座標

                

                if keys4[i][0]=="+" or keys4[i][0]=="-":  

                    

                    if len(coor_name)!=0:  ##如果沒有連結，但是運算符號為+或-，則從coor_name中，找一個主事者與連結，變成主事者接受者數量為2
                        if '=' in  keys4[0]:
                            keys1[i].append(coor_name[0][0])  ##主事者接受者數量為2
                        else:
                            coor_name.append([name,label,["none"],[]])  ##另存一個新的list
                    else:
                        coor_name.append([name,label,["none"],[]])  ##如果coor_name為空，則coor_name存入這個名稱跟位置
                        print("coor_name","存入",[name,label,["none"],[]])
                        

                else: ##如果運算符號為=
                    
                    coor_name.append([name,label,["none"],[]])  ##存入名稱跟位置
                    print("coor_name","存入",[name,label,["none"],[]])
                

            count+=1

            pre.clear()
            pre.append(label)   ##存入前一個標籤的位置
            
            print("存入前一個標籤的位置(pre):",pre)
            
            

        if len(keys1[i])==2:  ##主事者接受者數量為2

            print("主事者接受者數量為2")

            ##暫存名稱
            name1 = keys1[i][0]
            name2 = keys1[i][1]
            ##暫存名稱

            ##暫存位置
            label = num[i]+str(count)
            count+=1
            label2 = num[i]+str(count)
            count+=1
            ##暫存位置

            print("名單 : ",name1,"，",name2)
                       

            for x in coor_name:

                
                if x[0]==name1:     ##有找到一樣名稱的標籤，產生連結
                    print("產生連結，",name1)   
                    label = x[1]   ##現有位置為產生連結之標籤的位置
                    linked1 += 1
                    break
                elif x[0]==name2:     ##有找到一樣名稱的標籤，產生連結
                    print("產生連結，",name2)  
                    label2 = x[1]  ##現有位置為產生連結之標籤的位置
                    linked2 += 1
                    break
                    
            
                
            if linked1==linked2==0:  ##如果沒有連結，則從物品單位等做連結
                print("未產生連結，往下尋找")  
                for x in coor_item:

                    if x[0]==name1:     ##有找到一樣名稱的標籤，產生連結
                        
                        label = x[1]   ##現有位置為產生連結之標籤的位置
                        linked1 += 1
                        print('連結',name1)
                        break

                    elif x[0]==name2:     ##有找到一樣名稱的標籤，產生連結
                        
                        label2 = x[1]  ##現有位置為產生連結之標籤的位置
                        linked2 += 1
                        print('連結',name2)
                        break
                        


            if linked1==0:
                
                coor_name.append([name1,label,["none"],[]])     ##如果第一個名稱沒有連結，則存入名稱跟位置
                print("coor_name","存入",[name1,label,["none"],[]])
            if linked2==0:
                
                coor_name.append([name2,label2,["none"],[]])   ##存入名稱跟位置
                print("coor_name","存入",[name2,label2,["none"],[]])
                
            if linked1==0 and linked2==0:  ##如果linked1跟linked2都沒有被連結，則表示這句話的標籤沒有被連結，不做其他處理
                linked=0
            else:
                linked=1

            
                

            pre.clear()
            pre.append(label)   ##存入前一個標籤的位置
            pre.append(label2)  ##存入前一個標籤的位置
            
            print("前一個位置(pre):",pre)
            
        ##先處理主事者接受者

        #print()
        print("處理運算符號")

        ##處理+-=

        if len(keys1[i])==1:   ##主事者接受者有一個
            name = keys4[i][0]   ##暫存+-=

            
            label = num[i]+str(count)  ##暫存位置
            count+=1
            coor_plus.append([name,label,[pre[0]],[]])   ##存入名稱跟位置
            
            print("coor_plus","存入",[name,label,[pre[0]],[]])

            for x in coor_name:   ##如果前面的標籤有找到對應的位置，把目前位置存到coor_name的第三項(下一個位置)中

                if x[1]== pre[0]:
                    x[3].append(label)   
            
            pre.clear()
            pre.append(label)   ##存入前一個標籤的位置
            
            print("前一個位置(pre):",pre)

            
            
        elif len(keys1[i])==2:   ##主事者接受者有兩個
        
            print("主事者接受者有兩個")
            print("先處理主事者")
            
            name = keys4[i][0]   ##暫存+-=
            label = num[i]+str(count) ##暫存位置
            count+=1
            coor_plus.append([name,label,[pre[0]],[]])  ##存入名稱跟位置
            
            print("coor_plus","存入",[name,label,[pre[0]],[]])

            check1 = False ##確認是否有連結到
            check2 = False ##確認是否有連結到

            for x in coor_name:   ##從前面的標籤有找到對應的位置

                if x[1]== pre[0]:  ##如果位置跟pre標籤的一樣，則代表目前的標籤，上一個位置是為於這個標籤中
                    x[3].append(label)  ##這個標籤存入目前位置(等同於上一個標籤，存入下一個標籤位置)
                    check1 = True

            if check1==False:  ##如果沒有連結到，再從coor_item找連結

                for x in coor_item:   ##從前面的標籤有找到對應的位置

                    if x[1]== pre[0]:  ##如果位置跟pre標籤的一樣，則代表目前的標籤，上一個位置是為於這個標籤中
                        x[3].append(label)  ##這個標籤存入目前位置(等同於上一個標籤，存入下一個標籤位置)
                        check1 = True
                        
            
            ##根據前面的+-=，去判斷下一個+-=為何
            if name == "+":  ##如果前面為+，則接下來為-
                name = "-"  
            elif name == "-":  ##如果前面為-，則接下來為+
                name = "+"
            elif name=="=":
                name="="

            if name == "大" or name == "小":  ##如果+-=的詞性為大或小，則keys2,keys3，會存入此詞彙

                keys2[i].insert(0,keys5[i][0])
                keys3[i].insert(0,keys4[i][0])
                
                if name == "大":  ##如果前面為+，則接下來為-
                    name = "小"  
                elif name == "小":  ##如果前面為-，則接下來為+
                    name = "大"

            print("處理接受者")            

            label2 = num[i]+str(count)  ##暫存位置
            count+=1
            coor_plus.append([name,label2,[pre[1]],[]])  ##存入名稱跟位置
            
            print("coor_plus","存入",[name,label2,[pre[1]],[]])
            
            for x in coor_name:   ##如果前面的標籤有找到對應的位置，把目前位置存到前面的標籤中

                if x[1]== pre[1]:
                    x[3].append(label2)
                    check2 = True

            if check2==False:  ##如果沒有連結到，再從coor_item找連結

                for x in coor_item:   ##從前面的標籤有找到對應的位置

                    if x[1]== pre[1]:  ##如果位置跟pre標籤的一樣，則代表目前的標籤，上一個位置是為於這個標籤中
                        x[3].append(label2)  ##這個標籤存入目前位置(等同於上一個標籤，存入下一個標籤位置)
                        check2 = True
                        
            
            pre.clear()
            pre.append(label)
            pre.append(label2)

            print("前一個位置(pre):",pre)
        
        ##處理+-=

       
            
        ##處理物品數量單位
        
        print("往下處理(物品數量單位...等)")
        for j in range(len(keys2[i])):
            
            name = keys2[i][j]  #標籤名稱
            
            types = keys3[i][j] ##標籤種類

            label = num[i]+str(count)
            count+=1

            

            if len(pre)==1:  ##pre只有一個元素，表示只有一個主事者

                
                if linked!=0 or is_first==1:  ##如果有連結，則直接存入標籤



                    # if types=="opr" :   ##如果是+-=，則上一個的位置會變為主事者("沒用到")
                        

                    #     pre[0]=init_pos
                        
                    #     coor_item.append([name,label,[pre[0]],[]])   ##存入名稱跟位置

                        
                        
                    
                        
                        
                    # else:  #
                    coor_item.append([name,label,[pre[0]],[]])   ##存入名稱跟位置

                   

                    

                    

                elif types=="物品" or types=="單位" and linked==0:   ##如果沒有連結，而標籤種類為物品或單位

                    print("前面沒有連結")
                    
                    for k in coor_item:  ##從物品數量單位裡面找，是否有對應的項目
                        if k[0]==name:  ##如果有找到，表示產生連結
                            print("連結 : ",name)
                            linked+=1
                            label = k[1]   ##暫存位置
                            k[2].append(pre[0])
                            
                            
                            
                    if linked==0:
                        coor_item.append([name,label,[pre[0]],[]])   ##存入名稱跟位置
                        print("coor_item存入:",[name,label,[pre[0]],[]])
                        
                    
                else:
                    coor_item.append([name,label,[pre[0]],[]])   ##存入名稱跟位置
                    print("coor_item存入:",[name,label,[pre[0]],[]])

                
                print("處理後面位置")
                for x in coor_name:   ##如果前面的標籤有找到對應的位置(名稱)，把目前位置存到前面的標籤中
                    
                    if x[1]== pre[0]:
                        
                        x[3].append(label)
                        
                        print("coor_name:",x,"後面位置:",x[3])
                        
                for x in coor_plus:   ##如果前面的標籤有找到對應的位置(+-=)，把目前位置存到前面的標籤中

                    if x[1]== pre[0]:
                        x[3].append(label)
                        print("coor_plus:",x,"後面位置:",x[3])

                for x in coor_item:   ##如果前面的標籤有找到對應的位置(物品數量單位)，把目前位置存到前面的標籤中
                    
                    if x[1]== pre[0]:
                        x[3].append(label)
                        print("coor_item:",x,"後面位置:",x[3])
                        
                

                
                
                
            else:  ##如果主事者接受者有兩個

                if len(pre)!=0:
                
                    coor_item.append([name,label,[pre[0],pre[1]],[]])   ##存入名稱跟位置
                    
                    print("coor_item存入:",[name,label,[pre[0],pre[1]],[]])
                    
                    
                    print("處理後面位置")
                    for x in coor_plus:   ##如果前面的標籤有找到對應的位置(+-=)，把目前位置存到前面的標籤中

                        if x[1]== pre[0] or x[1]==pre[1]:
                            x[3].append(label)
                            print("coor_plus",x,"後面位置:",x[3])

                    for x in coor_item:   ##如果前面的標籤有找到對應的位置(物品數量單位)，把目前位置存到前面的標籤中
                        
                        if x[1]== pre[0] or x[1]==pre[1]:
 
                            x[3].append(label)
                            print("coor_item",x,"後面位置:",x[3])

                  
            pre.clear()
            pre.append(label)    ##存入前一個標籤的位置

            
    for x in coor_item:
        if len(x[3])==0:   ##如果下一個位置為空值(沒有存入位置)，表示已經到終點
            x[3].append("none")   ##存入none表示到終點
            
    
    print()
        
#########################把輸出資料寫入txt###########################################################################

    for x in coor_name:  ##把主事者接受者的list放到output_list
        output_list.append(x)

    for x in coor_plus:   ##把+-=的list放到output_list
        output_list.append(x)
    

    for x in coor_item:   ##把物品數量單位的list放到output_list
        output_list.append(x)

    if is_plus==False:  ##題目中有"比"這個詞，ex: 我比你多10元。=> 之後會做特殊計算
        output_list.append(["比較","a100",["a0"],["none"]])
    if is_remain == True :  ##題目中有["剩下","剩"]這些詞，ex: 我有100元，買1個橡皮擦，剩下80元。 => 做減法倒推計算
        output_list.append(["剩下","a101",["a0"],["none"]])
    if is_do ==True:   ##題目中有["做"]或 ["用" , "用掉"]這些詞，ex: 我有3顆蘋果，做果汁用掉2顆 => 解題目前沒用到
        output_list.append(["做","a102",["a0"],["none"]])
    
    if is_total ==True: ##題目有["共有","總共","共","一共","一共有"]這類詞，例如 : 遊樂園共有100人，男生有50人。女生有幾人?=> 做減法計算
        output_list.append(["總共","a103",["a0"],["none"]])
    if change_plus==True:   ##題目有["其中"]這類詞，例如 : 遊樂園有100人，其中男生有50人。女生有幾人?=> 做減法計算
        output_list.append(["其中","a104",["a0"],["none"]])
    
    if plus_one==True:  ##總共排多少人，計算時要+1。例如 : 我前面排了5人，後面排了3人。總共排了多少人?
        output_list.append(["總排","a105",["a0"],["none"]])

    if add_back==True:  ##題目有["不夠"]這類詞，例如 : 
        output_list.append(["不夠","a106",["a0"],["none"]])
    
    if no_have==True:
        output_list.append(["沒、沒有","a107",["a0"],["none"]])

############################把輸出資料寫入txt###########################################################################
        

    for i in output_list:   ##印出list內容
        #print(i)
        #print()
        
        f.write(str(i[0]))  ##寫入名稱
        f.write("\n")
        f.write(str(i[1]))  ##寫入位置
        f.write("\n")
        for x in i[2]:  ##寫入前一個標籤的位置
           f.write(str(x)+" ")
        f.write("\n")
        for x in i[3]:  ##寫入下一個標籤的位置
           f.write(str(x)+" ")
        f.write("\n")

    

    
    for k in output_list:  ##印出來
        print(k)

    
    
    f.close()
    
    
    
