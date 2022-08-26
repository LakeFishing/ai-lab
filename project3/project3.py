
## 建立語意網路 : 

## 1.把語意圖畫出來 : 會開啟一個小視窗，裡面有3個按鈕，點選按鈕，開啟新的小視窗來繪製語意圖
## 2.輸出txt檔案=> 給project4讀取、解題


##############################   project3.py使用之function   ##############################

## load_excel():  ##把project2輸出的excel檔讀進來

## new_page(pic_title) : 開啟新的視窗，繪製語意圖，會使用下面三個function來畫圖。 pic_title : 讓function判斷目前要處理的是哪一種語意圖

## show1(cv,sheet) 畫命題語意網路。  cv : 畫布(canvas)  sheet:excel裡面的sheet
## show2(cv,sheet) 畫句子語意網路。  cv : 畫布(canvas)  sheet:excel裡面的sheet
## show3(cv,sheet) 畫問題語意網路。  cv : 畫布(canvas)  sheet:excel裡面的sheet

## read3(sheet)  把讀進來的excel檔，進行處理，輸出txt檔。

## save_image(title,cv)   儲存圖片成jpg檔。

##############################   project3.py使用之function   ##############################

import sys

sys.dont_write_bytecode = True



##import 的部分#####################################################################

from tkinter import *   
from random import randint
import pyautogui  ##如果要在介面上存圖檔，使用這個來截圖存檔
from openpyxl import load_workbook

#import p3_read      ##讀取project2輸出的excel檔案並整理，接下來才能進行下面的步驟

from . import p3_show1     ##命題語意網路繪圖
from . import p3_show2     ##句子語意網路繪圖
from . import p3_show3     ##問題語意網路繪圖
from . import p3_output    ##把語意網路連結成問題語意網路格式，輸出成txt檔給project4用

# import project4.project4 as proj4

import os
import time

##import 的部分#####################################################################





##

def load_excel():  ##把project2輸出的excel檔讀進來
    

    wb = load_workbook(filename = 'proj2.xlsx')
    sheet = wb.active

    return sheet
    


        

def new_page(pic_title):   #新增視窗顯示圖片，pic_title為圖像種類名稱，如命題語意網路或句子語意網路或問題語意網路

    #sheet = load_excel()  ##讀取excel

    #a,b,c,d = read3(sheet)
    #calculate(a,b,c,d)

    ##讀取excel
    sheet = load_excel()  
    read3(sheet)  ##寫入txt
    
        
    
    top1 = Toplevel()##新增視窗
    top1.title(pic_title)##視窗的名稱
    
    cv = Canvas(top1,width=1200,height=500)   ##新增canvas，將圖片顯示在上面
    cv.pack(expand=True)

    
    if pic_title =="命題語意網路":
        
        final_pic = show1(cv,sheet)  ##畫圖，return出來的final_pic是最後要儲存的圖

    elif pic_title =="句子語意網路":
        
        final_pic = show2(cv,sheet)  ##畫圖，return出來的final_pic是最後要儲存的圖

    else:
        final_pic = show3(cv,sheet)  ##畫圖，return出來的final_pic是最後要儲存的圖

    ##按鈕是用來觸發儲存圖片的function，輸入變數有:圖片以及圖片類型
    
    b4 = Button(top1,text="儲存圖片",command=lambda: save_image(pic_title,cv))
    b4.pack()



def show1(cv,sheet):       ###問題語意網路


    p3_show1.draw(cv,sheet)
    
            
    
def show2(cv,sheet):     ##句子語意網路

    p3_show2.draw(cv,sheet) 

def show3(cv,sheet):     ##問題語意網路

    p3_show3.draw(cv,sheet)
    
def read3(sheet):   ##把語意網路寫入txt檔，解題用

    p3_output.read3(sheet)

    
def save_image(title,cv):  ##使用截圖，儲存成圖片jpg

    filename = ""
    cv.update()
    
    
    filename+=".jpg"
    filename = title+filename
    #image.save(filename)   ##儲存成圖片jpg檔，以亂數命名，會標示是哪一種類型之圖片，例如:命題語意網路12345678.jpg

    x = cv.winfo_rootx()
    y = cv.winfo_rooty()
    w =cv.winfo_width()
    h =cv.winfo_height()
    
    ##cv.delete("all")canvas清空
    
    pic = pyautogui.screenshot(region=(x,y,w,h))  ##截圖
    pic.save(filename)
    

oval_w = 60 #橢圓的寬
oval_h = 30   #橢圓的高


############這個區塊，如果執行UI.py時，需註解掉。平常自己測試的時候不用註解############

def test():
    root = Tk()  
    root.geometry("300x300+500+300")

    b1 = Button(root,text="命題語意網路",command=lambda: new_page("命題語意網路"))    ##新增命題語意網路視窗
    b1.pack()

    b2 = Button(root,text="句子語意網路",command=lambda: new_page("句子語意網路"))    ##新增句子語意網路視窗
    b2.pack()

    b3 = Button(root,text="問題語意網路",command=lambda: new_page("問題語意網路"))    ##新增問題語意網路視窗
    b3.pack()


    root.mainloop()  ##執行mainloop


############這個區塊，如果執行UI.py時，需註解掉。平常自己測試的時候不用註解############

def main():
    
    # folder = r'../output'
    # files = os.listdir(folder)
    # for file in files:
    #     if file.endswith('.xlsx'):
    #         cwd = os.getcwd()  # Get the current working directory (cwd)
    #         files = os.listdir(cwd)  # Get all the files in that directory
    #         print("Files in %r: %s" % (cwd, files))
    #         wb = load_workbook(filename= '../output/' + file)
    #         sheet = wb.active
    #         read3(sheet)
    #         ans = proj4.initial()
    #         print(ans)
    # wb = load_workbook(filename = '../output/proj2_5.xlsx')
    wb = load_workbook(filename = 'proj2.xlsx')
    sheet = wb.active
    read3(sheet)

if __name__ == '__main__':
    main()
# test()
# main()
